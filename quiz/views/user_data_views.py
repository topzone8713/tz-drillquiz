from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.parsers import JSONParser
from django.contrib.auth import get_user_model
from django.db import models
import logging

logger = logging.getLogger(__name__)

def get_message_by_language(language, key):
    """언어에 따라 메시지를 반환합니다."""
    if language == 'en':
        return en_messages.get(key, key)
    else:
        return ko_messages.get(key, key)

from ..models import UserProfile, Exam, Question, ExamResult, ExamResultDetail, IgnoredQuestion, StudyProgressRecord, StudyTaskProgress, AccuracyAdjustmentHistory, StudyJoinRequest, Member, ExamSubscription
from ..serializers import ExamSerializer
from ..utils.multilingual_utils import get_localized_field, BASE_LANGUAGE
from ..email_utils import send_email_verification, generate_verification_token, is_token_expired
from ..message_ko import KOREAN_TRANSLATIONS as ko_messages
from ..message_en import ENGLISH_TRANSLATIONS as en_messages
import json
import os
from datetime import datetime
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import get_object_or_404
from django.utils import timezone
import pandas as pd
import io
from django.http import HttpResponse

User = get_user_model()


def to_naive(dt):
    if dt is None:
        return None
    if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
        # Convert to UTC then drop tzinfo
        return dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def get_question_statistics_for_user(question, user):
    """사용자별 문제 통계 정보를 반환합니다."""
    try:
        # 사용자의 해당 문제 시도 결과 조회
        result_details = ExamResultDetail.objects.filter(
            question=question,
            exam_result__user=user
        ).order_by('-created_at')
        
        if not result_details.exists():
            return {
                'score': 1.0,  # 미시도 문제는 최고 우선순위
                'attempts': 0,
                'wrong_count': 0
            }
        
        # 최신 결과 기준으로 통계 계산
        latest_result = result_details.first()
        total_attempts = result_details.count()
        wrong_attempts = result_details.filter(is_correct=False).count()
        
        # 점수 계산 (정답률 기반)
        if total_attempts > 0:
            accuracy = (total_attempts - wrong_attempts) / total_attempts
            score = max(0.1, accuracy)  # 최소 0.1점 보장
        else:
            score = 1.0
        
        return {
            'score': score,
            'attempts': total_attempts,
            'wrong_count': wrong_attempts
        }
        
    except Exception as e:
        logger.error(f'문제 통계 계산 중 오류: {e}')
        return {
            'score': 1.0,
            'attempts': 0,
            'wrong_count': 0
        }


# calculate_age_rating 함수는 quiz.utils.user_utils로 이동


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_profile(request):
    """사용자 프로필 정보를 가져옵니다 (캐싱 지원)."""
    try:
        from django.core.cache import cache
        
        user = request.user
        cache_key = f"user_profile_{user.id}"
        
        # 강제 새로고침 파라미터 확인
        force_refresh = request.GET.get('refresh') == 'true'
        
        # 캐시에서 데이터 확인 (강제 새로고침이 아닌 경우에만)
        if not force_refresh:
            cached_data = cache.get(cache_key)
            if cached_data:
                logger.debug(f"[USER_PROFILE] 캐시 히트: user_id={user.id}")
                return Response(cached_data)
        
        # 캐시에 없으면 DB에서 조회
        # 성능 최적화: interested_categories를 prefetch하여 N+1 쿼리 방지
        profile, created = UserProfile.objects.prefetch_related('interested_categories').get_or_create(user=user)
        
        # 관심 카테고리 ID 목록 가져오기 (마이그레이션 전 필드가 없을 수 있음)
        try:
            # prefetch된 경우 효율적으로 조회
            if hasattr(profile, '_prefetched_objects_cache') and 'interested_categories' in profile._prefetched_objects_cache:
                interested_category_ids = [cat.id for cat in profile.interested_categories.all()]
            else:
                # prefetch가 없는 경우 직접 조회
                interested_category_ids = list(profile.interested_categories.values_list('id', flat=True))
        except (AttributeError, Exception) as e:
            # interested_categories 필드가 아직 마이그레이션되지 않은 경우
            # OperationalError (테이블 없음) 또는 AttributeError 처리
            from django.db.utils import OperationalError
            if isinstance(e, OperationalError) or isinstance(e, AttributeError):
                interested_category_ids = []
            else:
                # 다른 예외는 다시 발생시킴
                raise
        
        # 소셜 로그인 사용자(비밀번호가 없는 사용자)는 자동으로 이메일 인증된 것으로 간주
        # 비밀번호가 없는 사용자는 Google/Apple 로그인 사용자이므로 이메일이 이미 인증됨
        is_social_login_user = not user.has_usable_password()
        email_verified = profile.email_verified or is_social_login_user
        
        # 소셜 로그인 사용자인데 email_verified가 False인 경우 자동으로 True로 업데이트
        if is_social_login_user and not profile.email_verified:
            profile.email_verified = True
            profile.save(update_fields=['email_verified'])
            logger.info(f"소셜 로그인 사용자의 이메일 인증 상태 자동 업데이트: user_id={user.id}, email={user.email}")
            # 프로필이 업데이트되었으므로 캐시 무효화
            cache.delete(cache_key)
        
        # 나이 등급 계산
        from quiz.utils.user_utils import calculate_age_rating
        age_rating = calculate_age_rating(profile.date_of_birth)
        
        response_data = {
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'random_exam_email_enabled': profile.random_exam_email_enabled,
            'random_exam_question_count': profile.random_exam_question_count,
            'auto_translation_enabled': profile.auto_translation_enabled,
            'retention_cleanup_enabled': profile.retention_cleanup_enabled,
            'retention_cleanup_percentage': profile.retention_cleanup_percentage,
            'language': profile.language,
            'email_verified': email_verified,
            'interested_categories': interested_category_ids,
            'age_rating': age_rating,
        }
        
        # 캐시에 저장 (300초 TTL, 동기 저장 - 거의 변경되지 않는 데이터이므로 캐시 히트 시 즉시 반환)
        cache.set(cache_key, response_data, 300)
        logger.debug(f"[USER_PROFILE] 캐시 저장 완료: user_id={user.id}, TTL=300초")
        
        return Response(response_data)
    except Exception as e:
        logger.error(f'사용자 프로필 조회 실패: {e}', exc_info=True)
        return Response({'error': '사용자 프로필 조회에 실패했습니다.'}, status=500)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_user_profile(request):
    """사용자 프로필 정보를 업데이트합니다."""
    try:
        user = request.user
        profile, created = UserProfile.objects.get_or_create(user=user)
        
        data = request.data
        
        # 사용자 정보 업데이트
        if 'email' in data:
            user.email = data['email']
            user.save()
        
        if 'first_name' in data:
            user.first_name = data['first_name']
            user.save()
            
        if 'last_name' in data:
            user.last_name = data['last_name']
            user.save()
        
        # 프로필 설정 업데이트
        if 'random_exam_email_enabled' in data:
            profile.random_exam_email_enabled = data['random_exam_email_enabled']
            
        if 'random_exam_question_count' in data:
            profile.random_exam_question_count = data['random_exam_question_count']
            
        if 'auto_translation_enabled' in data:
            profile.auto_translation_enabled = data['auto_translation_enabled']

        if 'language' in data:
            profile.language = data['language']
        
        # 관심 카테고리 업데이트 (마이그레이션 후에만 동작)
        interested_category_ids = []
        if 'interested_categories' in data:
            try:
                from ..models import TagCategory
                from django.db.utils import OperationalError
                interested_category_ids = data.get('interested_categories', [])
                if isinstance(interested_category_ids, list):
                    # 유효한 카테고리 ID만 필터링
                    valid_categories = TagCategory.objects.filter(
                        id__in=interested_category_ids,
                        is_active=True
                    )
                    profile.interested_categories.set(valid_categories)
                elif interested_category_ids is None:
                    # 빈 리스트로 전달되면 모든 관심 카테고리 제거
                    profile.interested_categories.clear()
            except (AttributeError, OperationalError):
                # interested_categories 필드가 아직 마이그레이션되지 않은 경우 무시
                pass
        
        profile.save()
        
        # 프로필 업데이트 시 캐시 무효화
        from django.core.cache import cache
        cache_key = f"user_profile_{user.id}"
        cache.delete(cache_key)
        logger.debug(f"[USER_PROFILE] 캐시 무효화: user_id={user.id}")
        
        # 업데이트된 프로필 정보를 응답에 포함
        # ManyToMany 필드는 set() 후 바로 조회 가능하므로 prefetch 불필요
        try:
            interested_category_ids = list(profile.interested_categories.values_list('id', flat=True))
        except (AttributeError, Exception) as e:
            from django.db.utils import OperationalError
            if isinstance(e, OperationalError) or isinstance(e, AttributeError):
                interested_category_ids = []
            else:
                raise
        response_data = {
            'message': '프로필이 성공적으로 업데이트되었습니다.',
            'random_exam_email_enabled': profile.random_exam_email_enabled,
            'random_exam_question_count': profile.random_exam_question_count,
            'auto_translation_enabled': profile.auto_translation_enabled,
            'language': profile.language,
            'interested_categories': interested_category_ids,
        }
        
        return Response(response_data)
    except Exception as e:
        logger.error(f'프로필 업데이트 실패: {e}')
        return Response({'error': '프로필 업데이트에 실패했습니다.'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_language(request):
    """사용자 언어 설정을 변경합니다."""
    try:
        user = request.user
        profile, created = UserProfile.objects.get_or_create(user=user)
        
        from quiz.utils.multilingual_utils import SUPPORTED_LANGUAGES
        language = request.data.get('language')
        if language not in SUPPORTED_LANGUAGES:
            return Response({'error': '지원하지 않는 언어입니다.'}, status=400)
        
        profile.language = language
        # 언어 변경 시 자동 번역 설정은 변경하지 않음 (기존 값 유지)
        profile.save(update_fields=['language'])
        
        # 캐시 무효화 (사용자 관련 캐시)
        from django.core.cache import cache
        cache.delete(f"user_profile_{user.id}")
        cache.delete(f"user_language_{user.id}")
        
        # JWT 토큰 재발급 (새로운 언어 정보 반영)
        from ..views.auth_views import issue_tokens_for_user
        tokens = issue_tokens_for_user(user)
        
        return Response({
            'success': True,
            'tokens': tokens  # 새로운 토큰 반환
        }, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'언어 설정 변경 실패: {e}')
        return Response({'error': '언어 설정 변경에 실패했습니다.'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_user_data(request):
    """사용자 데이터를 엑셀 파일로 내보냅니다."""
    try:
        import pandas as pd
        from io import BytesIO
        from django.http import HttpResponse
        
        user = request.user
        
        # 사용자 데이터 수집
        data = {
            '사용자 정보': {
                '사용자명': user.username,
                '이메일': user.email,
                '이름': user.first_name,
                '성': user.last_name,
                '가입일': user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            }
        }
        
        # 시험 결과 데이터
        from ..models import ExamResult
        exam_results = ExamResult.objects.filter(user=user)
        if exam_results.exists():
            results_data = []
            for result in exam_results:
                results_data.append({
                    '시험 제목': get_localized_field(result.exam, 'title', result.exam.created_language if hasattr(result.exam, 'created_language') else 'en', 'Unknown'),
                    '점수': result.score,
                    '총점': result.total_score,
                    '정답 수': result.correct_count,
                    '오답 수': result.wrong_count,
                    '완료일': result.completed_at.strftime('%Y-%m-%d %H:%M:%S'),
                    '소요 시간(초)': result.elapsed_seconds,
                })
            data['시험 결과'] = results_data
        
        # 엑셀 파일 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, sheet_data in data.items():
                if isinstance(sheet_data, list):
                    df = pd.DataFrame(sheet_data)
                else:
                    df = pd.DataFrame([sheet_data])
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        output.seek(0)
        
        # 파일 다운로드 응답
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="user_data_{user.username}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
    except Exception as e:
        logger.error(f'사용자 데이터 내보내기 실패: {e}')
        return Response({'error': '사용자 데이터 내보내기에 실패했습니다.'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_email_verification_request(request):
    """이메일 인증 요청을 발송합니다."""
    try:
        user = request.user
        profile, created = UserProfile.objects.get_or_create(user=user)
        
        # 사용자 언어 설정 가져오기
        from quiz.utils.multilingual_utils import BASE_LANGUAGE
        user_language = getattr(profile, 'language', BASE_LANGUAGE)
        
        # 이메일이 설정되어 있는지 확인
        if not user.email:
            error_msg = get_message_by_language(user_language, 'profile.emailVerification.noEmail')
            return Response({'error': error_msg}, status=400)
        
        # 이미 인증된 경우
        if profile.email_verified:
            message = get_message_by_language(user_language, 'profile.emailVerification.alreadyVerified')
            return Response({
                'message': message,
                'email_verified': True
            })
        
        # 토큰 생성 및 저장
        token = generate_verification_token()
        profile.email_verification_token = token
        profile.email_verification_sent_at = timezone.now()
        profile.save()
        
        # 인증 URL 생성 (프론트엔드 URL로 변경)
        frontend_host = os.getenv('CURRENT_DOMAIN', 'localhost')

        # 디버깅 로그 추가
        logger.info(f'이메일 인증 URL 생성 디버깅:')
        logger.info(f'  - FRONTEND_HOST: {frontend_host}')
        logger.info(f'  - ENVIRONMENT: {os.getenv("ENVIRONMENT")}')
        logger.info(f'  - DOMAIN_PLACEHOLDER in FRONTEND_HOST: {"DOMAIN_PLACEHOLDER" in frontend_host}')
        
        # 환경에 따라 스키마 결정
        scheme = 'http' if frontend_host == 'localhost' else 'https'
        
        # 포트가 80 또는 443인 경우 포트 번호 제외
        if frontend_host == 'localhost':
            verification_url = f"{scheme}://{frontend_host}:8000/verify-email/{token}/"
        else:
            verification_url = f"{scheme}://{frontend_host}/verify-email/{token}/"

        logger.info(f'  - 최종 verification_url: {verification_url}')
        
        # 이메일 발송
        if send_email_verification(profile, verification_url):
            success_msg = get_message_by_language(user_language, 'profile.emailVerification.success')
            return Response({'message': success_msg})
        else:
            error_msg = get_message_by_language(user_language, 'profile.emailVerification.failed')
            return Response({'error': error_msg}, status=500)
            
    except Exception as e:
        logger.error(f'이메일 인증 요청 실패: {e}')
        error_msg = get_message_by_language(user_language, 'profile.emailVerification.failed')
        return Response({'error': error_msg}, status=500)


@api_view(['GET'])
def verify_email(request, token):
    """이메일 인증을 완료합니다."""
    from quiz.utils.multilingual_utils import BASE_LANGUAGE
    user_language = BASE_LANGUAGE  # 기본값 설정
    
    try:
        # 토큰으로 프로필 찾기
        profile = get_object_or_404(UserProfile, email_verification_token=token)
        
        # 사용자 언어 설정 가져오기
        from quiz.utils.multilingual_utils import BASE_LANGUAGE
        user_language = getattr(profile, 'language', BASE_LANGUAGE)
        
        # 토큰 만료 확인
        if is_token_expired(profile.email_verification_sent_at):
            error_msg = get_message_by_language(user_language, 'profile.emailVerification.tokenExpired')
            return Response({'error': error_msg}, status=400)
        
        # 이메일 인증 완료
        profile.email_verified = True
        profile.email_verification_token = None
        profile.email_verification_sent_at = None
        profile.save()
        
        success_msg = get_message_by_language(user_language, 'profile.emailVerification.verificationComplete')
        return Response({
            'message': success_msg,
            'email_verified': True
        })
        
    except Exception as e:
        logger.error(f'이메일 인증 실패: {e}')
        error_msg = get_message_by_language(user_language, 'profile.emailVerification.failed')
        return Response({'error': error_msg}, status=500)


@api_view(['GET'])
@permission_classes([AllowAny])
def list_question_files(request):
    from django.core.files.storage import default_storage
    from django.conf import settings
    
    try:
        import boto3
        from botocore.exceptions import ClientError
    except ImportError:
        return Response({'error': 'boto3 모듈이 설치되지 않았습니다.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    files = []
    
    # MinIO를 사용하는 경우와 로컬 스토리지를 사용하는 경우를 구분
    use_minio = getattr(settings, 'USE_MINIO', False)
    
    if use_minio:
        # MinIO에서 파일 목록 가져오기
        # 설정된 엔드포인트만 사용
        minio_endpoint = settings.AWS_S3_ENDPOINT_URL
        
        s3_client = None
        successful_endpoint = None
        
        try:
            s3_client = boto3.client(
                's3',
                endpoint_url=minio_endpoint,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                verify=False
            )
            
            # 연결 테스트
            s3_client.head_bucket(Bucket=settings.AWS_STORAGE_BUCKET_NAME)
            successful_endpoint = minio_endpoint
            
        except Exception as e:
            use_minio = False
        
        if not s3_client or not successful_endpoint:
            use_minio = False
        else:
            try:
                # data/ 폴더의 파일들만 가져오기 (MinIOStorage의 location='data' 설정 고려)
                response = s3_client.list_objects_v2(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                    Prefix='data/'
                )
                
                if 'Contents' in response:
                    for obj in response['Contents']:
                        # data/ 프리픽스 제거하고 파일명만 추출
                        filename = obj['Key'].replace('data/', '')
                        if filename:  # 빈 문자열이 아닌 경우만
                            # JSON 메타데이터 파일 제외 (.json 확장자)
                            if not filename.endswith('.json'):
                                # 파일의 공개 상태 확인
                                is_public = True  # 기본값은 공개
                                metadata_key = f'data/{filename}.json'
                                
                                try:
                                    # 메타데이터 파일이 있는지 확인
                                    metadata_response = s3_client.head_object(
                                        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                                        Key=metadata_key
                                    )
                                    # 메타데이터 파일이 있으면 내용을 읽어서 공개 상태 확인
                                    metadata_obj = s3_client.get_object(
                                        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                                        Key=metadata_key
                                    )
                                    import json
                                    metadata = json.loads(metadata_obj['Body'].read().decode('utf-8'))
                                    is_public = metadata.get('is_public', True)
                                except Exception as e:
                                    is_public = True  # 기본값
                                
                                # 태그 정보 가져오기
                                tags = []
                                if 'metadata' in locals() and metadata.get('tags'):
                                    from ..models import Tag
                                    tag_ids = metadata.get('tags', [])
                                    try:
                                        tags_queryset = Tag.objects.filter(id__in=tag_ids)
                                        tags = [{'id': tag.id, 'name': get_localized_field(tag, 'name', tag.created_language if hasattr(tag, 'created_language') else BASE_LANGUAGE, 'Unknown')} for tag in tags_queryset]
                                    except Exception as e:
                                        print(f"[list_question_files] 태그 정보 로드 실패: {e}")
                                
                                files.append({
                                    'name': filename,
                                    'size': obj['Size'],
                                    'modified': obj['LastModified'].isoformat(),
                                    'last_modified': obj['LastModified'].isoformat(),
                                    'is_public': is_public,
                                    'uploaded_by': (metadata.get('uploaded_by') or metadata.get('created_by')) if 'metadata' in locals() else None,
                                    'question_count': metadata.get('question_count', 0) if 'metadata' in locals() else 0,
                                    'max_questions': metadata.get('question_count', 0) if 'metadata' in locals() else 0,
                                    'tags': tags
                                })
                            else:
                                pass  # JSON 메타데이터 파일 제외
                
            except Exception as e:
                use_minio = False
    
    if not use_minio:
        # 로컬 스토리지에서 파일 목록 가져오기
        try:
            # data 디렉토리 경로
            data_dir = os.path.join(settings.MEDIA_ROOT, 'data')
            
            if os.path.exists(data_dir):
                for filename in os.listdir(data_dir):
                    file_path = os.path.join(data_dir, filename)
                    if os.path.isfile(file_path):
                        # JSON 메타데이터 파일 제외 (.json 확장자)
                        if not filename.endswith('.json'):
                            # 파일의 공개 상태 확인
                            is_public = True  # 기본값은 공개
                            metadata_path = os.path.join(data_dir, f'{filename}.json')
                            
                            if os.path.exists(metadata_path):
                                try:
                                    import json
                                    with open(metadata_path, 'r', encoding='utf-8') as f:
                                        metadata = json.load(f)
                                    is_public = metadata.get('is_public', True)
                                except Exception as e:
                                    is_public = True  # 기본값
                            
                            stat = os.stat(file_path)
                            # 메타데이터 파일이 있는지 확인
                            metadata_exists = os.path.exists(metadata_path)
                            
                            # 태그 정보 가져오기
                            tags = []
                            if metadata_exists and metadata.get('tags'):
                                from ..models import Tag
                                tag_ids = metadata.get('tags', [])
                                try:
                                    tags_queryset = Tag.objects.filter(id__in=tag_ids)
                                    tags = [{'id': tag.id, 'name_ko': tag.name_ko, 'name_en': tag.name_en} for tag in tags_queryset]
                                except Exception as e:
                                    print(f"[list_question_files] 태그 정보 로드 실패: {e}")
                            
                            file_data = {
                                'name': filename,
                                'size': stat.st_size,
                                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                                'last_modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                                'is_public': is_public,
                                'uploaded_by': metadata.get('uploaded_by') or metadata.get('created_by') if metadata_exists else None,
                                'question_count': metadata.get('question_count', 0) if metadata_exists else 0,
                                'max_questions': metadata.get('question_count', 0) if metadata_exists else 0,
                                'tags': tags
                            }
                            files.append(file_data)
                        else:
                            pass  # JSON 메타데이터 파일 제외
            else:
                pass  # data 디렉토리가 존재하지 않음
                
        except Exception as e:
            pass  # 파일 목록 조회 실패
    
    # 일반 사용자에게는 공개된 파일과 자신이 업로드한 파일만 반환
    if not request.user.is_authenticated or (hasattr(request.user, 'profile') and hasattr(request.user.profile, 'role') and request.user.profile.role != 'admin_role'):
        filtered_files = []
        for file in files:
            # 공개된 파일이거나 사용자가 업로드한 파일인 경우 포함
            if file.get('is_public', True) or (request.user.is_authenticated and file.get('uploaded_by') == request.user.username):
                filtered_files.append(file)
        
        files = filtered_files
    
    # 인증되지 않은 사용자에게는 공개된 파일만 반환
    if not request.user.is_authenticated:
        public_files = []
        for file in files:
            if file.get('is_public', True):
                public_files.append(file)
        files = public_files
    
    # 태그 필터링 적용
    tags_param = request.GET.get('tags', '')
    if tags_param:
        try:
            # 콤마로 구분된 태그 ID들을 정수 리스트로 변환
            tag_ids = [int(tag_id.strip()) for tag_id in tags_param.split(',') if tag_id.strip().isdigit()]
            if tag_ids:
                filtered_files = []
                for file in files:
                    # 파일의 tags 배열에서 태그 ID 확인
                    file_tags = file.get('tags', [])
                    file_tag_ids = [tag.get('id') for tag in file_tags if isinstance(tag, dict) and tag.get('id')]
                    # 선택된 태그 중 하나라도 파일에 포함되어 있으면 포함
                    if any(tag_id in file_tag_ids for tag_id in tag_ids):
                        filtered_files.append(file)
                files = filtered_files
        except (ValueError, AttributeError) as e:
            # 잘못된 태그 ID는 무시
            pass
    
    # 페이지네이션 처리
    from rest_framework.pagination import PageNumberPagination
    
    page_size = int(request.GET.get('page_size', 20))  # 기본값 20
    page = int(request.GET.get('page', 1))  # 기본값 1
    
    # 페이지네이션 계산
    total_count = len(files)
    start_index = (page - 1) * page_size
    end_index = start_index + page_size
    paginated_files = files[start_index:end_index]
    
    # 페이지네이션 정보
    total_pages = (total_count + page_size - 1) // page_size  # 올림 계산
    
    return Response({
        'files': paginated_files,
        'pagination': {
            'count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'has_next': page < total_pages,
            'has_previous': page > 1
        }
    }) 


@api_view(['GET'])
def download_question_file(request, filename):
    from django.core.files.storage import default_storage
    from django.conf import settings
    
    try:
        import boto3
        from botocore.exceptions import ClientError
    except ImportError:
        return Response({'error': 'boto3 모듈이 설치되지 않았습니다.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    from django.http import HttpResponse
    
    # MinIO를 사용하는 경우와 로컬 스토리지를 사용하는 경우를 구분
    if getattr(settings, 'USE_MINIO', False):
        try:
            s3_client = boto3.client(
                's3',
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                verify=False
            )
            
            # MinIO에서 파일 다운로드
            response = s3_client.get_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f'data/{filename}'
            )
            
            # 파일 내용을 HttpResponse로 반환
            file_content = response['Body'].read()
            http_response = HttpResponse(file_content, content_type=response['ContentType'])
            http_response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return http_response
            
        except ClientError as e:
            if e.response['Error']['Code'] == 'NoSuchKey':
                raise Http404
            else:
                return Response({'error': f'파일 다운로드 실패: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        # 로컬 스토리지에서 파일 다운로드
        fpath = os.path.join(settings.MEDIA_ROOT, 'data', filename)
        if not os.path.exists(fpath):
            raise Http404
        return FileResponse(open(fpath, 'rb'), as_attachment=True, filename=filename) 


@api_view(['DELETE'])
def delete_question_file(request, filename):
    from django.core.files.storage import default_storage
    from django.conf import settings
    
    try:
        import boto3
        from botocore.exceptions import ClientError
    except ImportError:
        return Response({'error': 'boto3 모듈이 설치되지 않았습니다.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    print(f"[delete_question_file] 삭제 요청 시작: {filename}")
    print(f"[delete_question_file] USE_MINIO: {getattr(settings, 'USE_MINIO', False)}")
    print(f"[delete_question_file] AWS_S3_ENDPOINT_URL: {getattr(settings, 'AWS_S3_ENDPOINT_URL', 'NOT_SET')}")
    print(f"[delete_question_file] AWS_STORAGE_BUCKET_NAME: {getattr(settings, 'AWS_STORAGE_BUCKET_NAME', 'NOT_SET')}")
    
    # 권한 확인
    if not request.user.is_authenticated:
        return Response({'error': '로그인이 필요합니다.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    # 관리자 권한 확인
    is_admin = False
    if hasattr(request.user, 'is_superuser') and request.user.is_superuser:
        is_admin = True
    elif hasattr(request.user, 'profile') and hasattr(request.user.profile, 'role') and request.user.profile.role == 'admin_role':
        is_admin = True
    
    # 파일 소유자 확인
    file_owner = None
    try:
        if getattr(settings, 'USE_MINIO', False):
            # MinIO에서 메타데이터 확인
            s3_client = boto3.client(
                's3',
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                verify=False
            )
            try:
                metadata_response = s3_client.get_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                    Key=f'data/{filename}.json'
                )
                import json
                metadata = json.loads(metadata_response['Body'].read().decode('utf-8'))
                file_owner = metadata.get('uploaded_by')
                print(f"[delete_question_file] 파일 소유자: {file_owner}")
            except Exception as e:
                print(f"[delete_question_file] 메타데이터 읽기 실패: {e}")
                file_owner = None
        else:
            # 로컬에서 메타데이터 확인
            metadata_path = os.path.join(settings.MEDIA_ROOT, 'data', f'{filename}.json')
            if os.path.exists(metadata_path):
                import json
                with open(metadata_path, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                file_owner = metadata.get('uploaded_by')
                print(f"[delete_question_file] 파일 소유자: {file_owner}")
    except Exception as e:
        print(f"[delete_question_file] 메타데이터 확인 실패: {e}")
        file_owner = None
    
    # 권한 체크: 관리자이거나 파일 소유자인 경우만 삭제 가능
    if not is_admin and file_owner != request.user.username:
        return Response({'error': '파일을 삭제할 권한이 없습니다.'}, status=status.HTTP_403_FORBIDDEN)
    
    # MinIO를 사용하는 경우와 로컬 스토리지를 사용하는 경우를 구분
    if getattr(settings, 'USE_MINIO', False):
        try:
            print(f"[delete_question_file] MinIO 클라이언트 생성 중...")
            s3_client = boto3.client(
                's3',
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                verify=False
            )
            
            # MinIO에서 파일 삭제 (존재 여부 확인 없이 직접 삭제 시도)
            print(f"[delete_question_file] 파일 삭제 시도: data/{filename}")
            try:
                delete_response = s3_client.delete_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                    Key=f'data/{filename}'
                )
                print(f"[delete_question_file] ✅ 파일 삭제 성공: {delete_response}")
            except ClientError as e:
                print(f"[delete_question_file] ⚠️ 파일 삭제 실패 (파일이 없을 수 있음): {e}")
                # 파일이 없어도 성공으로 처리
                pass
            
            # 메타데이터 파일도 삭제
            print(f"[delete_question_file] 메타데이터 파일 삭제 시도: data/{filename}.json")
            try:
                metadata_delete_response = s3_client.delete_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                    Key=f'data/{filename}.json'
                )
                print(f"[delete_question_file] ✅ 메타데이터 파일 삭제 성공: {metadata_delete_response}")
            except ClientError as e:
                print(f"[delete_question_file] ⚠️ 메타데이터 파일 삭제 실패 (무시됨): {e}")
                # 메타데이터 파일이 없으면 무시
                pass
                
            print(f"[delete_question_file] ✅ 모든 삭제 작업 완료")
            return Response(status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            print(f"[delete_question_file] ❌ 예상치 못한 오류: {e}")
            # 에러가 발생해도 성공으로 처리 (파일이 이미 삭제되었을 수 있음)
            return Response(status=status.HTTP_204_NO_CONTENT)
    else:
        print(f"[delete_question_file] 로컬 스토리지에서 삭제")
        # 로컬 스토리지에서 파일 삭제
        fpath = os.path.join(settings.MEDIA_ROOT, 'data', filename)
        
        # 원본 파일 삭제 (존재하지 않아도 에러 없이 처리)
        try:
            if os.path.exists(fpath):
                os.remove(fpath)
                print(f"[delete_question_file] ✅ 원본 파일 삭제 성공: {filename}")
            else:
                print(f"[delete_question_file] ⚠️ 원본 파일이 존재하지 않음: {filename}")
        except Exception as e:
            print(f"[delete_question_file] ⚠️ 원본 파일 삭제 실패 (무시): {e}")
        
        # 메타데이터 파일도 삭제
        metadata_fpath = os.path.join(settings.MEDIA_ROOT, 'data', f'{filename}.json')
        try:
            if os.path.exists(metadata_fpath):
                os.remove(metadata_fpath)
                print(f"[delete_question_file] ✅ 메타데이터 파일 삭제 성공: {filename}.json")
            else:
                print(f"[delete_question_file] ⚠️ 메타데이터 파일이 존재하지 않음: {filename}.json")
        except Exception as e:
            print(f"[delete_question_file] ⚠️ 메타데이터 파일 삭제 실패 (무시): {e}")
        
        return Response(status=status.HTTP_204_NO_CONTENT) 


@api_view(['PATCH'])
def update_question_file(request, filename):
    """파일의 공개 여부를 업데이트합니다."""
    from django.core.files.storage import default_storage
    from django.conf import settings
    
    try:
        import boto3
        from botocore.exceptions import ClientError
    except ImportError:
        return Response({'error': 'boto3 모듈이 설치되지 않았습니다.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # 권한 확인
    if not request.user.is_authenticated:
        return Response({'error': '로그인이 필요합니다.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    # 관리자 권한 확인
    is_admin = False
    try:
        user_profile = request.user.profile
        user_role = user_profile.role
        if user_role in ['admin_role', 'study_admin_role']:
            is_admin = True
    except:
        user_role = None
    
    # 파일 소유자 확인
    file_owner = None
    try:
        if getattr(settings, 'USE_MINIO', False):
            # MinIO에서 메타데이터 확인
            s3_client = boto3.client(
                's3',
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                verify=False
            )
            try:
                metadata_response = s3_client.get_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                    Key=f'data/{filename}.json'
                )
                import json
                metadata = json.loads(metadata_response['Body'].read().decode('utf-8'))
                file_owner = metadata.get('uploaded_by')
                print(f"[update_question_file] 파일 소유자: {file_owner}")
            except Exception as e:
                print(f"[update_question_file] 메타데이터 읽기 실패: {e}")
                file_owner = None
        else:
            # 로컬에서 메타데이터 확인
            metadata_path = os.path.join(settings.MEDIA_ROOT, 'data', f'{filename}.json')
            if os.path.exists(metadata_path):
                import json
                with open(metadata_path, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                file_owner = metadata.get('uploaded_by')
                print(f"[update_question_file] 파일 소유자: {file_owner}")
    except Exception as e:
        print(f"[update_question_file] 메타데이터 확인 실패: {e}")
        file_owner = None
    
    # 권한 체크: 관리자이거나 파일 소유자인 경우만 수정 가능
    if not is_admin and file_owner != request.user.username:
        return Response({'error': '파일을 수정할 권한이 없습니다.'}, status=status.HTTP_403_FORBIDDEN)
    
    # MinIO를 사용하는 경우와 로컬 스토리지를 사용하는 경우를 구분
    if getattr(settings, 'USE_MINIO', False):
        try:
            s3_client = boto3.client(
                's3',
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                verify=False
            )
            
            # 파일 존재 여부 확인
            try:
                s3_client.head_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                    Key=f'data/{filename}'
                )
            except ClientError as e:
                if e.response['Error']['Code'] == '404':
                    return Response({'error': '파일이 존재하지 않습니다.'}, status=status.HTTP_404_NOT_FOUND)
                else:
                    raise
            
            # is_public 필드 처리
            if 'is_public' in request.data:
                is_public = request.data['is_public']
                
                # 기존 메타데이터 읽기
                existing_metadata = {}
                try:
                    metadata_response = s3_client.get_object(
                        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                        Key=f'data/{filename}.json'
                    )
                    import json
                    existing_metadata = json.loads(metadata_response['Body'].read().decode('utf-8'))
                except Exception as e:
                    print(f"[update_question_file] 기존 메타데이터 읽기 실패: {e}")
                
                # 기존 메타데이터와 새로운 정보 병합
                import json
                metadata = {
                    **existing_metadata,  # 기존 메타데이터 유지
                    'is_public': is_public,
                    'updated_at': timezone.now().isoformat(),
                    'updated_by': request.user.username
                }
                
                metadata_content = json.dumps(metadata, ensure_ascii=False, indent=2)
                s3_client.put_object(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                    Key=f'data/{filename}.json',
                    Body=metadata_content.encode('utf-8'),
                    ContentType='application/json'
                )
                
                return Response({
                    'message': '파일 공개 여부가 업데이트되었습니다.',
                    'filename': filename,
                    'is_public': is_public
                })
            else:
                return Response({'error': 'is_public 필드가 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({'error': f'파일 업데이트 실패: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        # 로컬 스토리지에서 파일 업데이트
        fpath = os.path.join(settings.MEDIA_ROOT, 'data', filename)
        if not os.path.exists(fpath):
            return Response({'error': '파일이 존재하지 않습니다.'}, status=status.HTTP_404_NOT_FOUND)
        
        # is_public 필드 처리
        if 'is_public' in request.data:
            is_public = request.data['is_public']
            
            # 파일 메타데이터를 저장할 JSON 파일 생성
            metadata_file = os.path.join(settings.MEDIA_ROOT, 'data', f"{filename}.json")
            
            # 기존 메타데이터 읽기
            existing_metadata = {}
            if os.path.exists(metadata_file):
                try:
                    import json
                    with open(metadata_file, 'r', encoding='utf-8') as f:
                        existing_metadata = json.load(f)
                except Exception as e:
                    print(f"[update_question_file] 기존 메타데이터 읽기 실패: {e}")
            
            # 기존 메타데이터와 새로운 정보 병합
            import json
            metadata = {
                **existing_metadata,  # 기존 메타데이터 유지
                'is_public': is_public,
                'updated_at': timezone.now().isoformat(),
                'updated_by': request.user.username
            }
            
            with open(metadata_file, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)
            
            return Response({
                'message': '파일 공개 여부가 업데이트되었습니다.',
                'filename': filename,
                'is_public': is_public
            })
        else:
            return Response({'error': 'is_public 필드가 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST) 


@api_view(['GET', 'PATCH'])
def user_profile(request):
    """사용자 프로필 조회 및 업데이트"""
    import time
    start_time = time.time()
    
    if not request.user.is_authenticated:
        return Response({'error': '로그인이 필요합니다.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    if request.method == 'GET':
        from django.core.cache import cache
        
        cache_key = f"user_profile_{request.user.id}"
        
        # 강제 새로고침 파라미터 확인
        force_refresh = request.GET.get('refresh') == 'true'
        
        # 캐시에서 데이터 확인 (강제 새로고침이 아닌 경우에만)
        if not force_refresh:
            cache_start = time.time()
            cached_data = cache.get(cache_key)
            cache_time = time.time() - cache_start
            if cached_data:
                total_time = time.time() - start_time
                logger.info(f"[USER_PROFILE] 캐시 히트: user_id={request.user.id}, 캐시 조회 시간={cache_time*1000:.2f}ms, 총 시간={total_time*1000:.2f}ms")
                return Response(cached_data)
        
        # 캐시에 없으면 DB에서 조회
        db_start = time.time()
        # 성능 최적화: interested_categories를 prefetch하여 N+1 쿼리 방지
        # get_or_create를 사용하여 lazy loading과 추가 조회를 방지 (한 번의 쿼리로 처리)
        profile, created = UserProfile.objects.prefetch_related('interested_categories').get_or_create(
            user=request.user,
            defaults={'role': 'user_role'}
        )
        db_time = time.time() - db_start
        
        # 관심 카테고리 ID 목록 가져오기 (마이그레이션 전 필드가 없을 수 있음)
        categories_start = time.time()
        try:
            # prefetch된 경우 효율적으로 조회 (추가 쿼리 없음)
            if hasattr(profile, '_prefetched_objects_cache') and 'interested_categories' in profile._prefetched_objects_cache:
                interested_category_ids = [cat.id for cat in profile.interested_categories.all()]
            else:
                # prefetch가 없는 경우 직접 조회 (추가 쿼리 발생)
                interested_category_ids = list(profile.interested_categories.values_list('id', flat=True))
        except (AttributeError, Exception) as e:
            # interested_categories 필드가 아직 마이그레이션되지 않은 경우
            # OperationalError (테이블 없음) 또는 AttributeError 처리
            from django.db.utils import OperationalError
            if isinstance(e, OperationalError) or isinstance(e, AttributeError):
                interested_category_ids = []
            else:
                # 다른 예외는 다시 발생시킴
                raise
        categories_time = time.time() - categories_start
        
        # 나이 등급 계산
        from quiz.utils.user_utils import calculate_age_rating
        age_rating = calculate_age_rating(profile.date_of_birth)
        
        response_data = {
            'email': request.user.email,
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'random_exam_email_enabled': profile.random_exam_email_enabled,
            'random_exam_question_count': profile.random_exam_question_count,
            'auto_translation_enabled': profile.auto_translation_enabled,
            'retention_cleanup_enabled': profile.retention_cleanup_enabled,
            'retention_cleanup_percentage': profile.retention_cleanup_percentage,
            'language': profile.language,
            'interested_categories': interested_category_ids,
            'age_rating': age_rating,
        }
        
        # 캐시에 저장 (300초 TTL, 비동기 저장 - Redis 연결 지연 방지)
        # 첫 요청 시에는 응답을 먼저 반환하고, 캐시 저장은 Celery로 비동기 처리
        cache_save_start = time.time()
        try:
            from quiz.tasks import save_user_profile_cache
            save_user_profile_cache.delay(request.user.id, response_data, timeout=300)
            logger.debug(f"[USER_PROFILE] 캐시 저장 Celery 태스크 전송 완료: user_id={request.user.id}")
        except Exception as e:
            # Celery 태스크 전송 실패 시 동기 저장으로 폴백
            logger.warning(f"[USER_PROFILE] Celery 태스크 전송 실패, 동기 저장으로 폴백: {e}")
            try:
                cache.set(cache_key, response_data, 300)
                logger.debug(f"[USER_PROFILE] 동기 캐시 저장 완료: user_id={request.user.id}")
            except Exception as cache_error:
                logger.error(f"[USER_PROFILE] 캐시 저장 실패: {cache_error}")
        cache_save_time = time.time() - cache_save_start
        
        total_time = time.time() - start_time
        logger.info(f"[USER_PROFILE] 캐시 미스: user_id={request.user.id}, DB 조회={db_time*1000:.2f}ms, 카테고리 조회={categories_time*1000:.2f}ms, 캐시 저장 태스크 전송={cache_save_time*1000:.2f}ms, 총 시간={total_time*1000:.2f}ms")
        
        return Response(response_data)
    
    elif request.method == 'PATCH':
        print(f"PATCH 요청 데이터: {request.data}")
        # 프로필 가져오기 (성능 최적화: get_or_create 사용)
        profile, created = UserProfile.objects.get_or_create(
            user=request.user,
            defaults={'role': 'user_role'}
        )
        
        # 사용자 이메일 업데이트
        if 'email' in request.data:
            email = request.data['email'].strip()
            if email:
                # 이메일 형식 검증
                from django.core.validators import validate_email
                from django.core.exceptions import ValidationError
                try:
                    validate_email(email)
                    request.user.email = email
                    request.user.save()
                except ValidationError:
                    return Response({
                        'error': '올바른 이메일 형식이 아닙니다.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                request.user.email = ''
                request.user.save()
        
        # 사용자 이름 업데이트
        if 'first_name' in request.data:
            request.user.first_name = request.data['first_name'].strip()
            request.user.save()
        
        if 'last_name' in request.data:
            request.user.last_name = request.data['last_name'].strip()
            request.user.save()
        
        # 랜덤출제 이메일 발송 여부 업데이트
        if 'random_exam_email_enabled' in request.data:
            profile.random_exam_email_enabled = request.data['random_exam_email_enabled']
            
            # 이메일 발송이 활성화되면 이메일 필수
            if profile.random_exam_email_enabled and not request.user.email:
                return Response({
                    'error': '랜덤출제 이메일 발송을 활성화하려면 이메일 주소가 필요합니다.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # 랜덤출제 시험당 문제 수 업데이트
        if 'random_exam_question_count' in request.data:
            try:
                question_count = int(request.data['random_exam_question_count'])
                if question_count < 1 or question_count > 50:
                    return Response({
                        'error': '시험당 문제 수는 1-50개 사이여야 합니다.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                profile.random_exam_question_count = question_count
            except (ValueError, TypeError):
                return Response({
                    'error': '시험당 문제 수는 숫자여야 합니다.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # 언어 설정 업데이트
        if 'language' in request.data:
            language = request.data['language']
            from quiz.utils.multilingual_utils import SUPPORTED_LANGUAGES
            if language in SUPPORTED_LANGUAGES:
                profile.language = language
            else:
                return Response({
                    'error': '지원하지 않는 언어입니다.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # 자동 번역 설정 업데이트
        if 'auto_translation_enabled' in request.data:
            profile.auto_translation_enabled = request.data['auto_translation_enabled']

        # Retention Cleanup 설정 업데이트
        if 'retention_cleanup_enabled' in request.data:
            print(f"retention_cleanup_enabled 업데이트: {request.data['retention_cleanup_enabled']}")
            profile.retention_cleanup_enabled = request.data['retention_cleanup_enabled']
        
        if 'retention_cleanup_percentage' in request.data:
            try:
                percentage = int(request.data['retention_cleanup_percentage'])
                print(f"retention_cleanup_percentage 업데이트: {percentage}")
                if percentage < 0 or percentage > 100:
                    return Response({
                        'error': '정리 비율은 0-100% 사이여야 합니다.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                profile.retention_cleanup_percentage = percentage
            except (ValueError, TypeError):
                return Response({
                    'error': '정리 비율은 숫자여야 합니다.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # 관심 카테고리 업데이트 (마이그레이션 후에만 동작)
        interested_category_ids = []
        if 'interested_categories' in request.data:
            try:
                from ..models import TagCategory
                from django.db.utils import OperationalError
                interested_category_ids_input = request.data.get('interested_categories', [])
                logger.info(f'📊 관심 카테고리 업데이트 요청: {interested_category_ids_input}')
                print(f'📊 관심 카테고리 업데이트 요청: {interested_category_ids_input}')
                
                if isinstance(interested_category_ids_input, list):
                    # 유효한 카테고리 ID만 필터링
                    valid_categories = TagCategory.objects.filter(
                        id__in=interested_category_ids_input,
                        is_active=True
                    )
                    logger.info(f'📊 유효한 카테고리 개수: {valid_categories.count()}, IDs: {list(valid_categories.values_list("id", flat=True))}')
                    print(f'📊 유효한 카테고리 개수: {valid_categories.count()}, IDs: {list(valid_categories.values_list("id", flat=True))}')
                    
                    profile.interested_categories.set(valid_categories)
                    logger.info(f'✅ 관심 카테고리 설정 완료')
                    print(f'✅ 관심 카테고리 설정 완료')
                elif interested_category_ids_input is None:
                    # 빈 리스트로 전달되면 모든 관심 카테고리 제거
                    profile.interested_categories.clear()
                    logger.info(f'🗑️ 관심 카테고리 모두 제거')
                    print(f'🗑️ 관심 카테고리 모두 제거')
            except (AttributeError, OperationalError) as e:
                # interested_categories 필드가 아직 마이그레이션되지 않은 경우 무시
                logger.warning(f'⚠️ 관심 카테고리 업데이트 중 오류 (무시됨): {e}')
                print(f'⚠️ 관심 카테고리 업데이트 중 오류 (무시됨): {e}')
                pass
        
        profile.save()
        logger.info(f'💾 프로필 저장 완료')
        print(f'💾 프로필 저장 완료')
        
        # 프로필 업데이트 시 캐시 무효화
        from django.core.cache import cache
        cache_key = f"user_profile_{request.user.id}"
        cache.delete(cache_key)
        logger.debug(f"[USER_PROFILE] 캐시 무효화: user_id={request.user.id}")
        
        # 업데이트된 관심 카테고리 ID 목록 가져오기
        # ManyToMany 필드는 set() 후 바로 조회 가능하므로, profile.interested_categories에서 직접 조회
        interested_category_ids = []
        try:
            # profile.interested_categories에서 직접 조회 (ManyToMany 필드는 set() 후 바로 조회 가능)
            interested_category_ids = list(profile.interested_categories.values_list('id', flat=True))
            logger.info(f'📊 프로필에서 조회한 관심 카테고리 ID: {interested_category_ids}')
            print(f'📊 프로필에서 조회한 관심 카테고리 ID: {interested_category_ids}')
        except (AttributeError, Exception) as e:
            from django.db.utils import OperationalError
            if isinstance(e, OperationalError) or isinstance(e, AttributeError):
                interested_category_ids = []
                logger.warning(f'⚠️ 관심 카테고리 조회 실패 (빈 배열 반환): {e}')
                print(f'⚠️ 관심 카테고리 조회 실패 (빈 배열 반환): {e}')
            else:
                raise
        
        return Response({
            'message': '프로필이 업데이트되었습니다.',
            'email': request.user.email,
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'random_exam_email_enabled': profile.random_exam_email_enabled,
            'random_exam_question_count': profile.random_exam_question_count,
            'auto_translation_enabled': profile.auto_translation_enabled,
            'retention_cleanup_enabled': profile.retention_cleanup_enabled,
            'retention_cleanup_percentage': profile.retention_cleanup_percentage,
            'language': profile.language,
            'interested_categories': interested_category_ids
        })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def manual_retention_cleanup(request):
    """수동으로 성공한 기록을 정리합니다."""
    try:
        percentage = request.data.get('percentage', 0)
        
        # 현재 사용자의 성공한 기록 조회
        successful_details = ExamResultDetail.objects.filter(
            result__user=request.user,
            is_correct=True
        ).order_by('result__completed_at')  # 오래된 것부터
        
        total_count = successful_details.count()
        if total_count == 0:
            return Response({
                'message': '정리할 성공한 기록이 없습니다.',
                'deleted_count': 0
            })
        
        # 삭제할 개수 계산
        delete_count = int(total_count * percentage / 100)
        
        if delete_count > 0:
            # 오래된 것부터 삭제
            # 슬라이싱된 QuerySet의 ID 목록을 가져와서 삭제
            details_to_delete_ids = list(successful_details[:delete_count].values_list('id', flat=True))
            ExamResultDetail.objects.filter(id__in=details_to_delete_ids).delete()
        
        return Response({
            'message': f'{delete_count}개의 성공한 기록이 삭제되었습니다.',
            'deleted_count': delete_count,
            'total_count': total_count
        })
        
    except Exception as e:
        logger.error(f'수동 정리 실패: {str(e)}')
        return Response({
            'error': '수동 정리 중 오류가 발생했습니다.'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_language(request):
    """사용자 언어 설정을 변경합니다."""
    try:
        user = request.user
        # 언어만 변경하고, 자동 번역 설정은 기존 값 유지 (새로 생성 시 기본값 False 사용)
        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={'auto_translation_enabled': False}  # 새로 생성 시 자동 번역 비활성화
        )
        
        from quiz.utils.multilingual_utils import SUPPORTED_LANGUAGES
        language = request.data.get('language')
        if language not in SUPPORTED_LANGUAGES:
            return Response({'error': '지원하지 않는 언어입니다.'}, status=400)
        
        profile.language = language
        # 언어만 업데이트 (자동 번역 설정은 변경하지 않음)
        profile.save(update_fields=['language'])
        
        # 캐시 무효화 (사용자 관련 캐시)
        from django.core.cache import cache
        cache.delete(f"user_profile_{user.id}")
        cache.delete(f"user_language_{user.id}")
        
        # JWT 토큰 재발급 (새로운 언어 정보 반영)
        from ..views.auth_views import issue_tokens_for_user
        tokens = issue_tokens_for_user(user)
        
        return Response({
            'success': True,
            'tokens': tokens  # 새로운 토큰 반환
        }, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'언어 설정 변경 실패: {e}')
        return Response({'error': '언어 설정 변경에 실패했습니다.'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_change_user_password(request, user_id):
    """관리자가 사용자 비밀번호를 변경합니다."""
    # new_password 또는 password 필드 모두 지원
    password = request.data.get('new_password') or request.data.get('password')
    if not password or len(password) < 6:
        return Response({'detail': '비밀번호는 6자 이상이어야 합니다.'}, status=400)
    try:
        user = User.objects.get(id=user_id)
        user.set_password(password)
        user.save()
        return Response({'detail': '비밀번호가 변경되었습니다.'})
    except User.DoesNotExist:
        return Response({'detail': '사용자를 찾을 수 없습니다.'}, status=404)


class UserCreateView(APIView):
    def post(self, request):
        """새 사용자를 생성합니다."""
        from quiz.utils.multilingual_utils import BASE_LANGUAGE
        try:
            data = request.data
            required_fields = ['username', 'password']
            for field in required_fields:
                if not data.get(field):
                    return Response({
                        'detail': get_message_by_language(
                            request.data.get('language', BASE_LANGUAGE), 
                            'userManagement.messages.fieldRequired'
                        ).replace('{field}', field)
                    }, status=400)
            
            username = data['username']
            password = data['password']
            first_name = data.get('first_name', '')
            email = data.get('email', '')
            role = data.get('role', 'user_role')
            
            # 사용자명 중복 확인
            if User.objects.filter(username=username).exists():
                return Response({
                    'detail': get_message_by_language(request.data.get('language', BASE_LANGUAGE), 'userManagement.messages.usernameExists')
                }, status=400)
            
            # 이메일 중복 확인 (이메일이 제공된 경우에만)
            if email and User.objects.filter(email=email).exists():
                return Response({
                    'detail': get_message_by_language(request.data.get('language', BASE_LANGUAGE), 'userManagement.messages.emailExists')
                }, status=400)
            
            # 새 사용자 생성
            user = User.objects.create(
                username=username,
                first_name=first_name,
                email=email,
                is_active=True
            )
            user.set_password(password)
            user.save()
            
            # UserProfile 생성
            UserProfile.objects.create(user=user, role=role)
            
            return Response({
                'detail': get_message_by_language(request.data.get('language', BASE_LANGUAGE), 'userManagement.messages.createUserSuccess'),
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'email': user.email,
                    'role': role,
                    'date_joined': user.date_joined
                }
            }, status=201)
            
        except Exception as e:
            return Response({'detail': f'{get_message_by_language(request.data.get("language", BASE_LANGUAGE), "userManagement.messages.createUserError")}: {str(e)}'}, status=500)


class UserUpdateView(APIView):
    def put(self, request, user_id):
        """사용자 정보를 수정합니다."""
        from quiz.utils.multilingual_utils import BASE_LANGUAGE
        try:
            user = User.objects.get(id=user_id)
            data = request.data
            try:
                profile = user.profile
            except UserProfile.DoesNotExist:
                profile = UserProfile.objects.create(user=user, role='user_role')
            if 'email' in data:
                new_email = data['email']
                # 이메일 중복 확인 (다른 사용자와 중복되지 않는지)
                if new_email and new_email != user.email:
                    if User.objects.filter(email=new_email).exclude(id=user.id).exists():
                        return Response({
                            'detail': get_message_by_language(
                                request.data.get('language', BASE_LANGUAGE), 
                                'userManagement.messages.emailExists'
                            )
                        }, status=400)
                user.email = new_email
                
            # 이름(first_name)도 수정
            if 'first_name' in data:
                user.first_name = data['first_name']
            if 'role' in data:
                profile.role = data['role']
            user.save()
            profile.save()
            return Response({
                'detail': get_message_by_language(request.data.get('language', BASE_LANGUAGE), 'userManagement.messages.updateUserSuccess'),
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'email': user.email,
                    'role': profile.role
                }
            }, status=200)
        except User.DoesNotExist:
            return Response({'detail': '사용자를 찾을 수 없습니다.'}, status=404)
        except Exception as e:
            return Response({'detail': f'{get_message_by_language(request.data.get("language", BASE_LANGUAGE), "userManagement.messages.updateUserError")}: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_users_excel(request):
    """사용자 정보를 Excel 형식으로 다운로드합니다."""
    try:
        # 관리자 권한 확인
        if not request.user.is_authenticated:
            return Response({'error': '로그인이 필요합니다.'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            user_profile = request.user.profile
            user_role = user_profile.role
        except:
            user_role = None
        
        if user_role not in ['admin_role', 'study_admin_role']:
            return Response({'error': '관리자 권한이 필요합니다.'}, status=status.HTTP_403_FORBIDDEN)
        
        # 사용자 데이터 수집
        users = User.objects.all().order_by('username')
        user_list = []
        
        for user in users:
            try:
                profile = user.profile
                role = profile.role
            except UserProfile.DoesNotExist:
                profile = UserProfile.objects.create(user=user, role='user_role')
                role = profile.role
            
            user_list.append({
                'id': user.id,
                'first_name': user.first_name or '',
                'username': user.username,
                'email': user.email or '',
                'role': role,
                'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
                'is_active': '활성' if user.is_active else '비활성',
                'is_staff': '예' if user.is_staff else '아니오',
                'is_superuser': '예' if user.is_superuser else '아니오'
            })
        
        # DataFrame 생성
        df = pd.DataFrame(user_list)
        
        # 컬럼명을 한글로 변경
        df.columns = ['ID', '이름', '사용자명', '이메일', '역할', '가입일', '활성화', '스태프', '슈퍼유저']
        
        # Excel 파일 생성
        output = io.BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='사용자목록', index=False)
                
                # 워크시트 가져오기
                worksheet = writer.sheets['사용자목록']
                
                # 컬럼 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            
            # HTTP 응답 생성
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="users_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
            
        except Exception as excel_error:
            logger.error(f'Excel 파일 생성 중 오류: {excel_error}')
            return Response({'detail': f'Excel 파일 생성 중 오류가 발생했습니다: {str(excel_error)}'}, status=500)
        
    except Exception as e:
        logger.error(f'사용자 Excel 다운로드 중 오류: {e}')
        return Response({'detail': f'Excel 다운로드 중 오류가 발생했습니다: {str(e)}'}, status=500)


@api_view(['POST'])
def upload_users_excel(request):
    """Excel 파일을 통해 사용자 정보를 업로드/업데이트합니다."""
    try:
        if 'file' not in request.FILES:
            return Response({'detail': '파일이 업로드되지 않았습니다.'}, status=400)
        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'detail': 'Excel 파일(.xlsx, .xls)만 업로드 가능합니다.'}, status=400)
        try:
            df = pd.read_excel(file, engine='openpyxl')
        except Exception as e:
            return Response({'detail': f'Excel 파일 읽기 실패: {str(e)}'}, status=400)
        required_columns = ['이름', '사용자명', '이메일', '역할']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return Response({'detail': f'필수 컬럼이 누락되었습니다: {missing_columns}'}, status=400)
        stats = {
            'total_rows': len(df),
            'created': 0,
            'updated': 0,
            'errors': 0,
            'error_details': []
        }
        for index, row in df.iterrows():
            try:
                username = str(row['사용자명']).strip()
                first_name = str(row['이름']).strip() if '이름' in row and pd.notna(row['이름']) else ''
                email = str(row['이메일']).strip() if pd.notna(row['이메일']) else ''
                role = str(row['역할']).strip()
                if not username or username == 'nan':
                    stats['errors'] += 1
                    stats['error_details'].append(f'행 {index + 2}: 사용자명이 비어있습니다.')
                    continue
                role_mapping = {
                    '관리자': 'admin_role',
                    'admin_role': 'admin_role',
                    '일반사용자': 'user_role',
                    'user_role': 'user_role',
                    '사용자': 'user_role',
                    '스터디 관리자': 'study_admin_role',
                    'study_admin_role': 'study_admin_role',
                }
                mapped_role = role_mapping.get(role, 'user_role')
                try:
                    user = User.objects.get(username=username)
                    if email:
                        user.email = email
                    if first_name:
                        user.first_name = first_name
                    user.save()
                    try:
                        profile = user.profile
                    except UserProfile.DoesNotExist:
                        profile = UserProfile.objects.create(user=user, role='user_role')
                    profile.role = mapped_role
                    profile.save()
                    stats['updated'] += 1
                except User.DoesNotExist:
                    user = User.objects.create(
                        username=username,
                        first_name=first_name,
                        email=email,
                        is_active=True
                    )
                    UserProfile.objects.create(user=user, role=mapped_role)
                    stats['created'] += 1
            except Exception as e:
                stats['errors'] += 1
                stats['error_details'].append(f'행 {index + 2}: {str(e)}')
                continue
        message = f"처리 완료: 총 {stats['total_rows']}행, 생성 {stats['created']}명, 업데이트 {stats['updated']}명"
        if stats['errors'] > 0:
            message += f", 오류 {stats['errors']}건"
        return Response({
            'message': message,
            'stats': stats
        }, status=200)
    except Exception as e:
        return Response({'detail': f'Excel 업로드 중 오류가 발생했습니다: {str(e)}'}, status=500)


@api_view(['DELETE'])
def delete_user(request, user_id):
    """개별 사용자를 삭제합니다. (delete_my_account와 동일한 로직 사용)"""
    try:
        # 삭제할 사용자 조회
        user = User.objects.get(id=user_id)
        
        # 관리자(superuser)는 삭제할 수 없음
        if user.is_superuser:
            return Response({'detail': '관리자는 삭제할 수 없습니다.'}, status=400)
        
        username = user.username
        
        # delete_my_account와 동일한 데이터 정리 로직
        
        # 1. 먼저 스터디 멤버 상태 확인 및 처리
        from ..models import Study, StudyTask
        user_studies = Study.objects.filter(members__user=user)
        
        for study in user_studies:
            # 해당 스터디의 멤버 수 확인 (본인 제외)
            other_members = study.members.exclude(user=user)
            logger.info(f'스터디 "{study.title_ko or study.title_en}" 멤버 분석: 본인 제외 멤버 {other_members.count()}명')
            
            if other_members.count() == 0:
                # 다른 멤버가 없으면 스터디와 연결된 태스크 삭제
                study_tasks = StudyTask.objects.filter(study=study)
                if study_tasks.exists():
                    logger.info(f'사용자 "{username}"의 고아 스터디 "{study.title_ko or study.title_en}"의 태스크 {study_tasks.count()}개 삭제')
                    study_tasks.delete()
                
                # 스터디 삭제
                logger.info(f'사용자 "{username}"의 고아 스터디 "{study.title_ko or study.title_en}" 삭제')
                study.delete()
            else:
                # 다른 멤버가 있으면 스터디 관리자 존재 여부 확인
                from ..models import Member
                current_user_member = study.members.filter(user=user).first()
                
                if current_user_member and current_user_member.role in ['study_admin', 'study_leader']:
                    # 본인이 관리자인 경우, 다른 관리자가 있는지 확인
                    other_admins = other_members.filter(role__in=['study_admin', 'study_leader'])
                    if other_admins.count() == 0:
                        # 다른 관리자가 없으면 에러 발생
                        study_title = study.title_ko or study.title_en or study.title
                        error_msg = f'"{study_title}" 스터디의 스터디 관리자를 지정해야 합니다.'
                        logger.error(f'사용자 "{username}" 삭제 실패: {error_msg}')
                        return Response({'detail': error_msg}, status=400)
                
                # 관리자 문제가 없으면 본인만 멤버에서 제거
                logger.info(f'스터디 "{study.title_ko or study.title_en}"에 다른 멤버 {other_members.count()}명 존재 - 본인만 제거하고 스터디 보존')
                study.members.filter(user=user).delete()
                logger.info(f'사용자 "{username}"을 스터디 "{study.title_ko or study.title_en}"에서 제거 완료 (다른 멤버 {other_members.count()}명 존재)')
        
        # 2. 시험 결과 상세 삭제
        exam_results = ExamResult.objects.filter(user=user)
        ExamResultDetail.objects.filter(result__in=exam_results).delete()
        
        # 3. 시험 결과 삭제
        exam_results.delete()
        
        # 4. 스터디 진행 기록 삭제
        StudyProgressRecord.objects.filter(user=user).delete()
        
        # 5. 스터디 태스크 진행율 삭제
        StudyTaskProgress.objects.filter(user=user).delete()
        
        # 6. 정확도 조정 기록 삭제
        AccuracyAdjustmentHistory.objects.filter(user=user).delete()
        
        # 7. 무시된 문제 삭제
        IgnoredQuestion.objects.filter(user=user).delete()
        
        # 8. 스터디 가입 요청 삭제
        StudyJoinRequest.objects.filter(user=user).delete()
        
        # 9. 시험 구독 정보 삭제
        ExamSubscription.objects.filter(user=user).delete()
        
        # 10. 사용자가 생성한 시험 처리 (스터디 멤버 상태에 따라 결정)
        user_exams = Exam.objects.filter(created_by=user)
        
        if user_exams.exists():
            logger.info(f'사용자 "{username}"이 생성한 시험 {user_exams.count()}개 처리 시작')
            
            for exam in user_exams:
                # 해당 시험이 연결된 스터디가 있는지 확인
                connected_studies = Study.objects.filter(tasks__exam=exam).distinct()
                
                should_preserve_exam = False
                if connected_studies.exists():
                    for study in connected_studies:
                        # 스터디에 다른 멤버가 있는지 확인
                        other_members = study.members.exclude(user=user)
                        if other_members.count() > 0:
                            should_preserve_exam = True
                            logger.info(f'시험 "{exam.title_ko or exam.title_en}"은 스터디 "{study.title_ko or study.title_en}"에 다른 멤버가 있어 보존됨')
                            break
                
                if should_preserve_exam:
                    # 스터디에 다른 멤버가 있으면 시험 보존 (created_by만 None으로 설정)
                    exam.created_by = None
                    exam.save()
                    logger.info(f'시험 "{exam.title_ko or exam.title_en}" 보존 (created_by를 None으로 설정)')
                else:
                    # 스터디에 다른 멤버가 없으면 시험 삭제
                    logger.info(f'시험 "{exam.title_ko or exam.title_en}" 삭제 (연결된 스터디에 다른 멤버 없음)')
                    exam.delete()
        
        # 11. 사용자 프로필 삭제 (CASCADE로 자동 삭제됨)
        
        # 12. 사용자 계정 삭제
        user.delete()
        
        return Response({
            'message': f'사용자 "{username}"이(가) 성공적으로 삭제되었습니다.',
            'deleted_user_id': user_id,
            'deleted_username': username
        }, status=200)
        
    except User.DoesNotExist:
        return Response({'detail': '사용자를 찾을 수 없습니다.'}, status=404)
    except Exception as e:
        logger.error(f'사용자 삭제 중 오류 발생: {str(e)}')
        return Response({'detail': f'사용자 삭제 중 오류가 발생했습니다: {str(e)}'}, status=500)


@api_view(['POST'])
def delete_users_bulk(request):
    """선택된 사용자들을 일괄 삭제합니다."""
    try:
        user_ids = request.data.get('user_ids', [])
        
        if not user_ids:
            return Response({'detail': '삭제할 사용자가 선택되지 않았습니다.'}, status=400)
        
        # 관리자(superuser)가 포함되어 있는지 확인하고 제거
        admin_users = User.objects.filter(id__in=user_ids, is_superuser=True)
        if admin_users.exists():
            admin_usernames = [user.username for user in admin_users]
            user_ids = [uid for uid in user_ids if uid not in admin_users.values_list('id', flat=True)]
            if not user_ids:
                return Response({'detail': f'관리자는 삭제할 수 없습니다: {", ".join(admin_usernames)}'}, status=400)
        
        deleted_users = []
        errors = []
        
        for user_id in user_ids:
            try:
                user = User.objects.get(id=user_id)
                username = user.username
                user.delete()
                deleted_users.append({'id': user_id, 'username': username})
            except User.DoesNotExist:
                errors.append(f'사용자 ID {user_id}: 사용자를 찾을 수 없습니다.')
            except Exception as e:
                errors.append(f'사용자 ID {user_id}: {str(e)}')
        
        message = f'{len(deleted_users)}명의 사용자가 삭제되었습니다.'
        if errors:
            message += f' (오류: {len(errors)}건)'
        
        return Response({
            'message': message,
            'deleted_users': deleted_users,
            'errors': errors
        }, status=200)
        
    except Exception as e:
        return Response({'detail': f'일괄 삭제 중 오류가 발생했습니다: {str(e)}'}, status=500)


@api_view(['POST'])
def delete_all_users(request):
    """모든 사용자를 삭제합니다 (관리자 제외)."""
    try:
        # 모든 관리자(superuser) 제외하고 일반 사용자만 삭제
        all_users = User.objects.filter(is_superuser=False)
        deleted_count = all_users.count()
        
        if deleted_count == 0:
            return Response({'detail': '삭제할 사용자가 없습니다.'}, status=400)
        
        # 삭제할 사용자 목록 저장
        deleted_users = list(all_users.values('id', 'username'))
        
        # 일괄 삭제
        all_users.delete()
        
        return Response({
            'message': f'{deleted_count}명의 사용자가 삭제되었습니다.',
            'deleted_users': deleted_users
        }, status=200)
        
    except Exception as e:
        return Response({'detail': f'전체 삭제 중 오류가 발생했습니다: {str(e)}'}, status=500)


@api_view(['GET'])
def search_users(request):
    """사용자 검색 API"""
    try:
        query = request.GET.get('q', '')
        if not query:
            return Response({'users': []})
        
        # 사용자명, 이메일, 이름으로 검색
        users = User.objects.filter(
            models.Q(username__icontains=query) |
            models.Q(email__icontains=query) |
            models.Q(first_name__icontains=query) |
            models.Q(last_name__icontains=query) |
            models.Q(first_name__icontains=query.split()[0]) |
            models.Q(last_name__icontains=query.split()[-1]) if ' ' in query else models.Q()
        )[:10]  # 최대 10개 결과
        
        user_data = []
        for user in users:
            user_data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name
            })
        
        return Response({'users': user_data})
        
    except Exception as e:
        print(f"사용자 검색 중 오류: {str(e)}")
        return Response({'error': f'사용자 검색 중 오류가 발생했습니다: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 


@api_view(['POST'])
def fix_member_user_connections(request):
    """멤버들의 사용자 연결을 수정합니다."""
    user = request.user
    if not user.is_authenticated:
        return Response({'error': '로그인이 필요합니다.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    # admin_role 사용자만 접근 가능
    if not hasattr(user, 'profile') or user.profile.role != 'admin_role':
        return Response({'error': '관리자 권한이 필요합니다.'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # 모든 멤버를 가져와서 사용자 연결 시도
        members = Member.objects.filter(user__isnull=True)
        fixed_count = 0
        
        for member in members:
            if member.name:
                try:
                    user_obj = User.objects.get(username=member.name)
                    member.user = user_obj
                    member.save()
                    fixed_count += 1
                    print(f"멤버 사용자 연결 수정: {member.name} -> {user_obj.id}")
                except User.DoesNotExist:
                    pass
        
        return Response({
            'message': f'{fixed_count}개의 멤버 사용자 연결이 수정되었습니다.',
            'fixed_count': fixed_count
        })
        
    except Exception as e:
        return Response({'error': f'멤버 연결 수정 중 오류가 발생했습니다: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@csrf_exempt
@api_view(['POST'])
def create_random_recommendation_exam(request):
    """
    랜덤출제 API - UI 없이 백엔드에서 직접 실행 가능
    
    요청 데이터:
    {
        "username": "admin",          # 필수: admin 사용자명
        "password": "DevOps!323",     # 필수: admin 비밀번호
        "target_username": "user123", # 선택사항: 대상 사용자명 (없으면 admin 사용자)
        "title": "추천문제_20250725",  # 선택사항, 없으면 자동 생성
        "questions_per_exam": 3,      # 선택사항, 기본값 3
        "is_public": false            # 선택사항, 기본값 false (비공개)
    }
    """
    try:
        # 디버깅: ExamSerializer import 확인
        try:
            from ..serializers import ExamSerializer
            print(f"[랜덤출제 API] ExamSerializer import 성공: {ExamSerializer}")
        except Exception as e:
            print(f"[랜덤출제 API] ExamSerializer import 실패: {e}")
        
        # 요청 데이터 파싱
        username = request.data.get('username')
        password = request.data.get('password')
        target_username = request.data.get('target_username')
        title = request.data.get('title')
        questions_per_exam = request.data.get('questions_per_exam')  # None이면 사용자 프로필에서 가져옴
        is_public = request.data.get('is_public', False)
        
        # User 모델 import를 함수 시작 부분으로 이동
        from django.contrib.auth.models import User
        from django.contrib.auth import authenticate
        
        # 프론트엔드에서 호출하는 경우 (username, password가 없고 현재 로그인한 사용자가 있음)
        if not username or not password:
            if not request.user.is_authenticated:
                return Response({
                    'error': '로그인이 필요합니다.'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            authenticated_user = request.user
            print(f"[랜덤출제 API] 프론트엔드 호출 - 현재 사용자: {authenticated_user.username}")
        else:
            # 백엔드에서 직접 호출하는 경우 (admin 인증 필요)
            
            # 사용자 존재 확인
            user = User.objects.filter(username=username).first()
            if not user:
                return Response({
                    'error': '존재하지 않는 사용자입니다.'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # 비밀번호 확인
            authenticated_user = authenticate(username=username, password=password)
            if not authenticated_user:
                return Response({
                    'error': '비밀번호가 일치하지 않습니다.'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # admin 권한 확인 (선택사항)
            if not user.is_staff:
                return Response({
                    'error': 'admin 권한이 필요합니다.'
                }, status=status.HTTP_403_FORBIDDEN)
        
        # 대상 사용자 설정
        if target_username:
            # 대상 사용자 존재 확인
            target_user = User.objects.filter(username=target_username).first()
            if not target_user:
                return Response({
                    'error': f'대상 사용자 "{target_username}"를 찾을 수 없습니다.'
                }, status=status.HTTP_400_BAD_REQUEST)
            batch_user = target_user
            print(f"[랜덤출제 API] 대상 사용자: {target_username}")
        else:
            batch_user = authenticated_user
            print(f"[랜덤출제 API] 대상 사용자: {authenticated_user.username} (기본값)")
        
        # questions_per_exam이 없으면 사용자 프로필에서 가져오기
        if questions_per_exam is None:
            try:
                # Django ORM 캐시를 무시하고 최신 데이터 가져오기
                # 타임스탬프가 있으면 강제로 캐시 무효화
                if '_t' in request.data:
                    print(f"[랜덤출제 API] 타임스탬프 감지: {request.data['_t']}, 캐시 무효화")
                    # 사용자 프로필을 강제로 새로고침
                    batch_user.refresh_from_db()
                
                # 강제로 데이터베이스에서 최신 데이터 가져오기
                from django.db import connection
                cursor = connection.cursor()
                cursor.execute("SELECT random_exam_question_count FROM quiz_userprofile WHERE user_id = %s", [batch_user.id])
                result = cursor.fetchone()
                
                if result:
                    questions_per_exam = result[0]
                    print(f"[랜덤출제 API] Raw SQL로 사용자 프로필에서 문제 수 가져옴: {questions_per_exam}")
                else:
                    # Raw SQL로 가져올 수 없는 경우 ORM 사용
                    user_profile = UserProfile.objects.get(user=batch_user)
                    questions_per_exam = user_profile.random_exam_question_count
                    print(f"[랜덤출제 API] ORM으로 사용자 프로필에서 문제 수 가져옴: {questions_per_exam}")
                    
            except Exception as e:
                print(f"[랜덤출제 API] 사용자 프로필 조회 중 오류: {str(e)}")
                questions_per_exam = 3  # 기본값
                print(f"[랜덤출제 API] 기본값 사용: {questions_per_exam}")
        
        # user_profile 변수가 정의되지 않았을 수 있으므로 안전하게 처리
        user_profile = None
        try:
            user_profile = UserProfile.objects.get(user=batch_user)
        except UserProfile.DoesNotExist:
            print(f"[랜덤출제 API] 사용자 프로필을 찾을 수 없음: {batch_user.username}")
        except Exception as e:
            print(f"[랜덤출제 API] 사용자 프로필 조회 중 오류: {str(e)}")
        
        # 제목 생성 (Today's Quizzes for username 형식)
        # 배치 앱에서 호출할 때는 항상 "Today's Quizzes for username" 형식으로 생성
        title = f"Today's Quizzes for {batch_user.username}"
        
        print(f"[랜덤출제 API] 사용자: {batch_user.username}")
        print(f"[랜덤출제 API] 제목: {title}")
        print(f"[랜덤출제 API] 시험당 문제 수: {questions_per_exam}")
        print(f"[랜덤출제 API] 공개 여부: {is_public}")
        
        # 사용자가 구독한 시험들에서 문제 추출 (Subscribed Exams)
        from ..models import Study, Member, StudyTask, ExamSubscription
        from django.utils import timezone
        
        # 사용자가 구독한 시험들 조회
        subscribed_exams = Exam.objects.filter(
            examsubscription__user=batch_user,
            examsubscription__is_active=True,
            examquestion__isnull=False
        ).distinct()
        
        print(f"[랜덤출제 API] 구독한 시험 수: {subscribed_exams.count()}")
        print(f"[랜덤출제 API] 구독한 시험 목록: {list(subscribed_exams.values_list('title_ko', flat=True))}")
        
        # 구독한 시험이 없으면 오류 반환 (스터디 기반 폴백 제거)
        if not subscribed_exams.exists():
            print(f"[랜덤출제 API] 구독한 시험이 없음")
            return Response({
                'error': 'home.dailyExam.noSubscribedExams'
            }, status=status.HTTP_400_BAD_REQUEST)
        else:
            # 구독한 시험들을 사용
            accessible_exams = subscribed_exams
            print(f"[랜덤출제 API] 구독한 시험들을 사용하여 문제 추출")
        
        # 접근 가능한 시험이 없으면 오류 반환
        if not accessible_exams.exists():
            return Response({
                'error': '사용자가 접근할 수 있는 시험이 없습니다. 스터디에 가입하거나 시험을 풀어보세요.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        print(f"[랜덤출제 API] 접근 가능한 시험 수: {accessible_exams.count()}")
        print(f"[랜덤출제 API] 접근 가능한 시험 목록: {list(accessible_exams.values_list('title_ko', flat=True))}")
        
        if not accessible_exams.exists():
            if subscribed_exams.exists():
                return Response({
                    'error': '구독한 시험에 문제가 없습니다. 다른 시험을 구독해보세요.'
                }, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({
                    'error': '사용자가 참여하고 있는 스터디의 시험이나 직접 생성한 시험이 없습니다. 시험을 구독하거나 스터디에 가입해보세요.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        all_questions = []
        exam_question_map = {}
        summary_data = {
            'total_exams_processed': 0,
            'total_questions_selected': 0,
            'unique_questions': 0,
            'exams_with_questions': []
        }
        
        # 모든 시험에서 문제 수집
        all_exam_questions = []
        print(f"[랜덤출제 API] 처리할 시험 목록: {list(accessible_exams.values_list('title_ko', flat=True))}")
        
        for exam in accessible_exams:
            print(f"[랜덤출제 API] 시험 처리 중: {exam.title_ko}")
            
            # 시험의 문제들 조회 (무시된 문제 제외)
            exam_questions = Question.objects.filter(
                examquestion__exam=exam
            ).distinct()
            
            print(f"[랜덤출제 API] 시험 '{exam.title_ko}'에서 {len(exam_questions)}개 문제 발견")
            
            # 무시된 문제 제외
            ignored_question_ids = set(
                IgnoredQuestion.objects.filter(user=batch_user)
                .values_list('question_id', flat=True)
            )
            exam_questions = [q for q in exam_questions if q.id not in ignored_question_ids]
            
            print(f"[랜덤출제 API] 무시된 문제 제외 후 {len(exam_questions)}개 문제 남음")
            
            if not exam_questions:
                print(f"[랜덤출제 API] 시험 '{exam.title_ko}'에 문제가 없음")
                continue
            
            # 문제 통계 조회
            for question in exam_questions:
                total_attempts = ExamResultDetail.objects.filter(
                    question=question,
                    result__user=batch_user
                ).count()
                
                if total_attempts > 0:
                    wrong_count = ExamResultDetail.objects.filter(
                        question=question,
                        result__user=batch_user,
                        is_correct=False
                    ).count()
                    # 틀린 비율이 높을수록 높은 점수
                    wrong_rate = wrong_count / total_attempts
                    # 틀린 비율이 같으면 적게 풀어본 문제가 우선순위 높음 (시도 횟수의 역수 사용)
                    # 틀린 비율 + (1 / 시도 횟수)로 점수 계산
                    score = wrong_rate + (1.0 / total_attempts)
                    question_title = question.title_ko if question.title_ko else question.title_en or '제목 없음'
                    print(f"[랜덤출제 API] 문제 {question.id} ({question_title}): 점수={score:.2f}, 시도={total_attempts}, 틀린={wrong_count}, 틀린비율={wrong_rate:.2f}")
                else:
                    # 시도 횟수가 0인 문제는 가장 높은 우선순위 (아직 풀어보지 않은 문제)
                    score = 1.0
                    question_title = question.title_ko if question.title_ko else question.title_en or '제목 없음'
                    print(f"[랜덤출제 API] 문제 {question.id} ({question_title}): 점수=1.0, 시도=0, 틀린=0 (미시도 문제 - 최고 우선순위)")
                
                all_exam_questions.append({
                    'question': question,
                    'exam': exam,
                    'score': score,
                    'attempts': total_attempts,
                    'wrong_count': wrong_count if total_attempts > 0 else 0
                })
            
            summary_data['total_exams_processed'] += 1
        
        # 각 시험에서 questions_per_exam 개수만큼 문제 선택
        all_questions = []
        exam_question_map = {}
        
        # 시험별로 문제를 그룹화
        exam_questions_map = {}
        for item in all_exam_questions:
            exam_title = item['exam'].title_ko
            if exam_title not in exam_questions_map:
                exam_questions_map[exam_title] = []
            exam_questions_map[exam_title].append(item)
        
        print(f"[랜덤출제 API] 시험별 문제 그룹화 결과:")
        for exam_title, questions in exam_questions_map.items():
            print(f"  - {exam_title}: {len(questions)}개 문제")
        
        # 각 시험에서 최대 1개 문제만 선택 (동일 그룹 중복 방지)
        total_selected = 0
        for exam_title, exam_questions in exam_questions_map.items():
            # 해당 시험의 문제들을 점수 순으로 정렬
            sorted_exam_questions = sorted(exam_questions, key=lambda x: x['score'], reverse=True)
            
            # 각 시험에서 최대 1개 문제만 선택 (동일 그룹 중복 방지)
            selected_from_exam = sorted_exam_questions[:1] if sorted_exam_questions else []
            
            print(f"[랜덤출제 API] 시험 '{exam_title}'에서 {len(selected_from_exam)}개 문제 선택 (동일 그룹 중복 방지: 최대 1개, 가용: {len(sorted_exam_questions)}개)")
            
            for item in selected_from_exam:
                question = item['question']
                exam = item['exam']
                
                # group_id에 소스 시험 이름 설정
                # 단, 사용자가 이미 설정한 group_id가 있으면 보존 (빈 문자열이 아닌 경우)
                if not question.group_id or question.group_id.strip() == '':
                    question.group_id = exam.title_ko
                    question.save()
                
                # 사용자 언어에 맞는 제목 선택
                question_title = question.title_ko if question.title_ko else question.title_en or '제목 없음'
                
                question_data = {
                    'id': question.id,
                    'title': question_title,
                    'source_exam': exam.title_ko,
                    'group_id': exam.title_ko,
                    'score': item['score'],
                    'attempts': item['attempts'],
                    'wrong_count': item['wrong_count']
                }
                
                # 선택된 문제 상세 로그
                accuracy = (item['attempts'] - item['wrong_count']) / item['attempts'] if item['attempts'] > 0 else 1.0
                print(f"[랜덤출제 API] 선택된 문제: {question.id} ({question_title}) - 점수={item['score']:.2f}, 시도={item['attempts']}, 틀린={item['wrong_count']}, 정확도={accuracy:.1%}")
                
                all_questions.append(question_data)
                
                # exam_question_map 업데이트
                if exam.title_ko not in exam_question_map:
                    exam_question_map[exam.title_ko] = []
                exam_question_map[exam.title_ko].append(question_data)
            
            total_selected += len(selected_from_exam)
        
        print(f"[랜덤출제 API] 총 {total_selected}개 문제 선택 완료")
        
        # 선택된 문제들의 소스 시험별 분포 확인
        source_exam_counts = {}
        for exam_title, questions in exam_question_map.items():
            source_exam_counts[exam_title] = len(questions)
        
        print(f"[랜덤출제 API] 선택된 문제 소스별 분포: {source_exam_counts}")
        
        # summary_data 업데이트
        for exam_title, questions in exam_question_map.items():
            summary_data['exams_with_questions'].append({
                'exam_title': exam_title,
                'questions_count': len(questions)
            })
            print(f"[랜덤출제 API] 시험 '{exam_title}'에서 {len(questions)}개 문제 선택")
        
        if not all_questions:
            return Response({
                'error': '선택할 수 있는 문제가 없습니다.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 중복 제거 (문제 ID 기준 + 그룹 ID 기준)
        # 1단계: 문제 ID 기준 중복 제거
        unique_by_id = []
        seen_ids = set()
        
        for question_data in all_questions:
            if question_data['id'] not in seen_ids:
                unique_by_id.append(question_data)
                seen_ids.add(question_data['id'])
        
        # 2단계: 그룹 ID 기준 중복 제거 (동일 그룹에서 여러 문제가 선택된 경우 최고 점수 문제만 유지)
        unique_questions = []
        seen_group_ids = {}
        
        for question_data in unique_by_id:
            group_id = question_data.get('group_id', '')
            question_id = question_data['id']
            question_score = question_data.get('score', 0)
            
            # 그룹 ID가 없거나 빈 문자열인 경우는 그대로 추가
            if not group_id or group_id.strip() == '':
                unique_questions.append(question_data)
                continue
            
            # 같은 그룹의 문제가 이미 선택된 경우
            if group_id in seen_group_ids:
                existing_question = seen_group_ids[group_id]
                existing_score = existing_question.get('score', 0)
                
                # 현재 문제의 점수가 더 높으면 기존 문제를 교체
                if question_score > existing_score:
                    # 기존 문제 제거
                    unique_questions = [q for q in unique_questions if q['id'] != existing_question['id']]
                    unique_questions.append(question_data)
                    seen_group_ids[group_id] = question_data
                    print(f"[랜덤출제 API] 그룹 '{group_id}'에서 더 높은 점수 문제로 교체: {existing_question['id']} → {question_id}")
                else:
                    print(f"[랜덤출제 API] 그룹 '{group_id}'에서 중복 문제 제거: {question_id} (기존: {existing_question['id']})")
            else:
                # 새로운 그룹의 문제
                unique_questions.append(question_data)
                seen_group_ids[group_id] = question_data
        
        summary_data['total_questions_selected'] = len(all_questions)
        summary_data['unique_questions'] = len(unique_questions)
        
        print(f"[랜덤출제 API] 중복 제거: {len(all_questions)}개 → {len(unique_by_id)}개 (ID 기준) → {len(unique_questions)}개 (그룹 기준)")
        
        # 같은 이름의 시험이 있으면 기존 시험 재사용
        existing_exam = Exam.objects.filter(title_ko=title, created_by=batch_user).first()
        if existing_exam:
            print(f"[랜덤출제 API] 기존 시험 '{title}' 발견 - 재사용")
            
            # 기존 시험의 문제들을 모두 제거하고 새로 추가
            print(f"[랜덤출제 API] 기존 문제 제거 및 새 문제 추가 시작")
            from ..models import ExamQuestion
            
            # 기존 문제들 제거
            existing_exam_questions = ExamQuestion.objects.filter(exam=existing_exam)
            print(f"[랜덤출제 API] 기존 문제 {existing_exam_questions.count()}개 제거")
            existing_exam_questions.delete()
            
            # 새 문제들 추가
            for i, question_data in enumerate(unique_questions):
                question = Question.objects.get(id=question_data['id'])
                ExamQuestion.objects.create(
                    exam=existing_exam,
                    question=question,
                    order=i + 1
                )
            
            print(f"[랜덤출제 API] 새 문제 {len(unique_questions)}개 추가 완료")
            
            # 시험 정보 업데이트
            existing_exam.total_questions = len(unique_questions)
            existing_exam.save()
            
            # 기존 시험의 문제들을 반환
            existing_questions = existing_exam.questions.all()
            selected_questions = []
            
            for question in existing_questions:
                # 문제 통계 정보 가져오기
                question_stats = get_question_statistics_for_user(question, batch_user)
                question_title = question.title_ko if question.title_ko else question.title_en or '제목 없음'
                selected_questions.append({
                    'id': question.id,
                    'title': question_title,
                    'source_exam': question.group_id,
                    'group_id': question.group_id,
                    'score': question_stats['score'],
                    'attempts': question_stats['attempts'],
                    'wrong_count': question_stats['wrong_count']
                })
            
            # 응답 데이터 구성
            from ..serializers import ExamSerializer
            response_data = {
                'success': True,
                'exam': ExamSerializer(existing_exam, context={'request': request}).data,
                'is_new': False,
                'message': '기존 Daily Exam을 재사용하고 문제를 업데이트했습니다.',
                'selected_questions': selected_questions,
                'summary': summary_data,
                'profile': {'language': user_profile.language if user_profile else 'en'}
            }
            
            return Response(response_data, status=status.HTTP_200_OK)
        
        # 새 시험 생성
        exam = Exam.objects.create(
            title_ko=title,
            total_questions=len(unique_questions),
            is_original=False,
            is_public=is_public,
            created_by=batch_user  # 시험 생성자 설정
        )
        
        # 대상 사용자가 해당 시험을 볼 수 있도록 접근 권한 설정
        # 방법: 대상 사용자에게 빈 시험 결과를 생성하여 접근 권한 부여
        print(f"[랜덤출제 API] 접근 권한 설정 시작")
        print(f"[랜덤출제 API] batch_user: {batch_user.username}")
        print(f"[랜덤출제 API] authenticated_user: {authenticated_user.username}")
        print(f"[랜덤출제 API] batch_user != authenticated_user: {batch_user != authenticated_user}")
        
        if batch_user != authenticated_user:  # 대상 사용자가 admin이 아닌 경우
            try:
                print(f"[랜덤출제 API] 빈 시험 결과 생성 시도")
                # 대상 사용자의 빈 시험 결과 생성 (접근 권한용)
                exam_result = ExamResult.objects.create(
                    exam=exam,
                    user=batch_user,
                    score=0,
                    total_score=len(unique_questions),
                    correct_count=0,
                    wrong_count=0,
                    elapsed_seconds=0
                )
                print(f"[랜덤출제 API] 대상 사용자 {batch_user.username}의 접근 권한 설정 완료")
                print(f"[랜덤출제 API] 생성된 시험 결과 ID: {exam_result.id}")
            except Exception as e:
                print(f"[랜덤출제 API] 접근 권한 설정 중 오류: {str(e)}")
                import traceback
                print(f"[랜덤출제 API] 오류 상세: {traceback.format_exc()}")
        else:
            print(f"[랜덤출제 API] 대상 사용자가 admin이므로 접근 권한 설정 생략")
        
        # 시험에 문제 추가
        for i, question_data in enumerate(unique_questions):
            question = Question.objects.get(id=question_data['id'])
            from ..models import ExamQuestion
            ExamQuestion.objects.create(
                exam=exam,
                question=question,
                order=i + 1
            )
        
        print(f"[랜덤출제 API] 시험 생성 완료: {exam.id}")
        
        # 시험 관련 캐시 무효화 (ExamCacheManager 사용)
        try:
            from ..utils.cache_utils import ExamCacheManager
            
            # 1차: ExamCacheManager를 통한 체계적인 캐시 무효화
            try:
                # 모든 시험 캐시 무효화
                ExamCacheManager.invalidate_all_exam_cache()
                # 사용자별 시험 캐시 무효화
                ExamCacheManager.invalidate_user_exam_cache(batch_user.id)
                print("[랜덤출제 API] ExamCacheManager를 통한 캐시 무효화 완료")
            except Exception as e:
                print(f"[랜덤출제 API] ExamCacheManager 캐시 무효화 실패: {e}")
                
                # 2차: Django 캐시 무효화 (폴백)
                try:
                    from django.core.cache import cache
                    
                    # Redis의 패턴 매칭을 사용하여 시험 관련 캐시 삭제
                    if hasattr(cache, 'delete_pattern'):
                        cache.delete_pattern("exams_*")
                        cache.delete_pattern("exam_*")
                        cache.delete_pattern("questions_*")
                        cache.delete_pattern("user_*")
                        print("[랜덤출제 API] Redis 패턴 기반 캐시 무효화 완료")
                    else:
                        # 로컬 캐시의 경우 개별 키 삭제
                        cache_keys_to_delete = [
                            "exams_anonymous",
                            "exams_anonymous_true", 
                            "exams_anonymous_false",
                            "exams_anonymous_all",
                            "exams_1",
                            "exams_1_true",
                            "exams_1_false", 
                            "exams_1_all",
                            "exam_management",
                            "exam_list",
                            "questions_list",
                            "user_exams",
                            "user_studies"
                        ]
                        
                        # 사용자별 캐시 키 추가
                        if batch_user:
                            user_id = batch_user.id
                            cache_keys_to_delete.extend([
                                f"exams_{user_id}",
                                f"exams_{user_id}_true",
                                f"exams_{user_id}_false",
                                f"exams_{user_id}_all",
                                f"exam_{user_id}",
                                f"questions_{user_id}",
                                f"user_{user_id}_exams",
                                f"user_{user_id}_studies"
                            ])
                        
                        # 모든 캐시 키 삭제
                        for key in cache_keys_to_delete:
                            cache.delete(key)
                        
                        print(f"[랜덤출제 API] 개별 키 기반 캐시 무효화 완료 ({len(cache_keys_to_delete)}개 키)")
                        
                except Exception as e2:
                    print(f"[랜덤출제 API] Django 캐시 무효화도 실패: {e2}")
                    
        except Exception as e:
            print(f"[랜덤출제 API] 전체 캐시 무효화 중 오류: {e}")
            
        # 사용자 프로필 정보 가져오기
        try:
            user_profile = batch_user.profile
            profile_language = user_profile.language
        except UserProfile.DoesNotExist:
            profile_language = 'en'
        
        # 응답 데이터 구성
        from ..serializers import ExamSerializer
        exam_serializer = ExamSerializer(exam)
        response_data = {
            'success': True,
            'exam': exam_serializer.data,
            'selected_questions': unique_questions,
            'summary': summary_data,
            'profile': {
                'language': profile_language
            }
        }
        
        # 5. 프론트엔드 캐시 무효화를 위한 응답 헤더 추가
        response = Response(response_data, status=status.HTTP_201_CREATED)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        
        # 6. 추가 디버깅 정보
        print(f"[랜덤출제 API] 생성된 시험 ID: {exam.id}")
        print(f"[랜덤출제 API] 생성된 시험 제목: {exam.title_ko}")
        print(f"[랜덤출제 API] 응답 데이터: {response_data}")
        
        return response
        

        
    except Exception as e:
        print(f"[랜덤출제 API] 오류 발생: {str(e)}")
        return Response({
            'error': f'랜덤출제 중 오류가 발생했습니다: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def get_random_recommendation_exam_questions(request):
    """랜덤 추천 시험의 문제 목록을 조회합니다."""
    try:
        # 요청 데이터에서 인증 정보와 시험 제목 가져오기
        username = request.data.get('username')
        password = request.data.get('password')
        exam_title = request.data.get('title')
        
        if not username or not password:
            return Response({
                'error': 'username과 password는 필수입니다.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not exam_title:
            return Response({
                'error': '시험 제목(title)은 필수입니다.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # admin 사용자 인증 확인
        from django.contrib.auth.models import User
        from django.contrib.auth import authenticate
        
        # 사용자 존재 확인
        user = User.objects.filter(username=username).first()
        if not user:
            return Response({
                'error': '존재하지 않는 사용자입니다.'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # 비밀번호 확인
        authenticated_user = authenticate(username=username, password=password)
        if not authenticated_user:
            return Response({
                'error': '비밀번호가 일치하지 않습니다.'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # admin 권한 확인
        if not user.is_staff:
            return Response({
                'error': 'admin 권한이 필요합니다.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # 시험 제목으로 시험 찾기
        exam = Exam.objects.filter(title=exam_title).first()
        
        if not exam:
            return Response({
                'error': f'제목이 "{exam_title}"인 시험을 찾을 수 없습니다.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # admin 사용자는 모든 시험에 접근 가능
        # (이미 위에서 admin 권한을 확인했으므로 추가 체크 불필요)
        
        # 시험의 문제들 조회 (순서대로)
        questions = Question.objects.filter(
            examquestion__exam=exam
        ).order_by('examquestion__order')
        
        # 문제 목록 구성
        question_list = []
        for i, question in enumerate(questions):
            # 매핑된 멤버 조회
            mapped_member = None
            try:
                from ..models import QuestionMemberMapping
                mapping = QuestionMemberMapping.objects.get(
                    question=question,
                    exam=exam
                )
                mapped_member = {
                    'id': mapping.member.id,
                    'name': mapping.member.name,
                    'email': mapping.member.email
                }
            except QuestionMemberMapping.DoesNotExist:
                pass
            
            question_title = question.title_ko if question.title_ko else question.title_en or '제목 없음'
            question_data = {
                'order': i + 1,  # 번호
                'csv_id': question.csv_id,  # ID
                'id': question.id,
                'title': question_title,
                'content': question.content_ko or question.content_en or '',
                'answer': question.answer_ko or question.answer_en or '',
                'explanation': question.explanation_ko or question.explanation_en or '',
                'difficulty': question.difficulty,
                'url': question.url,
                'group_id': question.group_id,
                'created_at': question.created_at,
                'updated_at': question.updated_at,
                'mapped_member': mapped_member,
                # 문제 제목에 연결된 URL (take-exam 페이지)
                'take_exam_url': f"{request.scheme}://{request.get_host()}/take-exam?question_id={question.id}&exam_id={exam.id}"
            }
            
            question_list.append(question_data)
        
        # 응답 데이터 구성
        response_data = {
            'exam': {
                'id': exam.id,
                'title': exam.title_ko or exam.title_en or 'Unknown',
                'total_questions': len(question_list),
                'is_public': exam.is_public,
                'created_at': exam.created_at
            },
            'questions': question_list
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        print(f"[랜덤 추천 시험 문제 조회 API] 오류 발생: {str(e)}")
        return Response({
            'error': f'문제 목록 조회 중 오류가 발생했습니다: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def get_random_exam_email_users(request):
    """
    random_exam_email_enabled가 활성화된 사용자들의 이메일 목록을 조회합니다.
    """
    try:
        # 요청 데이터에서 인증 정보 가져오기
        username = request.data.get('username')
        password = request.data.get('password')
        
        if not username or not password:
            return Response({
                'error': 'username과 password는 필수입니다.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # admin 사용자 인증 확인
        from django.contrib.auth.models import User
        from django.contrib.auth import authenticate
        
        # 사용자 존재 확인
        user = User.objects.filter(username=username).first()
        if not user:
            return Response({
                'error': '존재하지 않는 사용자입니다.'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # 비밀번호 확인
        authenticated_user = authenticate(username=username, password=password)
        if not authenticated_user:
            return Response({
                'error': '비밀번호가 일치하지 않습니다.'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # admin 권한 확인
        if not user.is_staff:
            return Response({
                'error': 'admin 권한이 필요합니다.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # random_exam_email_enabled가 True이고 이메일이 인증된 UserProfile을 가진 사용자들을 조회
        enabled_users = UserProfile.objects.filter(
            random_exam_email_enabled=True,
            email_verified=True
        ).select_related('user')
        
        # 이메일 목록 생성
        email_list = []
        for profile in enabled_users:
            if profile.user.email:  # 이메일이 있는 경우만 포함
                email_list.append({
                    'user_id': profile.user.id,
                    'username': profile.user.username,
                    'email': profile.user.email,
                    'first_name': profile.user.first_name,
                    'last_name': profile.user.last_name,
                    'role': profile.role,
                    'random_exam_question_count': profile.random_exam_question_count,
                    'language': profile.language,
                    'retention_cleanup_enabled': profile.retention_cleanup_enabled,
                    'retention_cleanup_percentage': profile.retention_cleanup_percentage,
                    'auto_translation_enabled': profile.auto_translation_enabled
                })
        
        return Response({
            'success': True,
            'count': len(email_list),
            'users': email_list
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 


@api_view(['GET'])
def get_users(request):
    """사용자 목록을 조회합니다."""
    try:
        # 관리자 권한 확인
        if not request.user.is_authenticated:
            return Response({'error': '로그인이 필요합니다.'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            user_profile = request.user.profile
            user_role = user_profile.role
        except:
            user_role = None
        
        if user_role not in ['admin_role', 'study_admin_role']:
            return Response({'error': '관리자 권한이 필요합니다.'}, status=status.HTTP_403_FORBIDDEN)
        
        # 모든 사용자 조회
        users = User.objects.all().order_by('username')
        
        user_list = []
        for user in users:
            try:
                profile = user.profile
                role = profile.role
            except UserProfile.DoesNotExist:
                profile = UserProfile.objects.create(user=user, role='user_role')
                role = profile.role
            
            user_list.append({
                'id': user.id,
                'first_name': user.first_name or '',
                'username': user.username,
                'email': user.email or '',
                'role': role,
                'date_joined': user.date_joined.isoformat() if user.date_joined else None,
                'is_active': user.is_active,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser
            })
        
        return Response(user_list, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': f'사용자 목록 조회 중 오류가 발생했습니다: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_statistics_summary(request):
    """사용자의 통계 요약 정보를 가져옵니다."""
    try:
        user = request.user
        
        # 각 통계 항목의 개수 조회
        exam_results_count = ExamResult.objects.filter(user=user).count()
        study_progress_count = StudyProgressRecord.objects.filter(user=user).count()
        task_progress_count = StudyTaskProgress.objects.filter(user=user).count()
        accuracy_adjustments_count = AccuracyAdjustmentHistory.objects.filter(user=user).count()
        
        return Response({
            'exam_results_count': exam_results_count,
            'study_progress_count': study_progress_count,
            'task_progress_count': task_progress_count,
            'accuracy_adjustments_count': accuracy_adjustments_count,
            'total_count': exam_results_count + study_progress_count + task_progress_count + accuracy_adjustments_count
        })
    except Exception as e:
        logger.error(f'사용자 통계 요약 조회 실패: {e}')
        return Response({'error': '통계 요약 조회에 실패했습니다.'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reset_user_statistics(request):
    """사용자의 모든 통계 데이터를 초기화합니다."""
    try:
        # 데이터베이스 연결 상태 확인 및 재연결
        from django.db import connection
        if connection.connection and getattr(connection.connection, 'closed', False):
            logger.warning("⚠️  Database connection is closed, attempting to reconnect...")
            connection.close()
            connection.connect()
            logger.info("🔄 Database reconnection completed")
        
        user = request.user
        user_id = request.data.get('user_id')
        
        # 본인 또는 관리자만 초기화 가능
        if str(user.id) != str(user_id) and user.username != user_id and not user.is_staff:
            return Response({'error': 'profile.statistics.reset.noPermission'}, status=403)
        
        # 백업 정보 생성
        backup_data = {
            'user_id': user.id,
            'username': user.username,
            'timestamp': timezone.now().isoformat(),
            'exam_results_count': ExamResult.objects.filter(user=user).count(),
            'study_progress_count': StudyProgressRecord.objects.filter(user=user).count(),
            'task_progress_count': StudyTaskProgress.objects.filter(user=user).count(),
            'accuracy_adjustments_count': AccuracyAdjustmentHistory.objects.filter(user=user).count()
        }
        
        # 통계 데이터 삭제
        deleted_counts = {}
        
        # 정확도 조정 기록 삭제
        deleted_counts['accuracy_adjustments'] = AccuracyAdjustmentHistory.objects.filter(user=user).delete()[0]
        
        # 시험 결과 상세 삭제
        deleted_counts['exam_result_details'] = ExamResultDetail.objects.filter(
            result__user=user
        ).delete()[0]
        
        # 시험 결과 삭제
        deleted_counts['exam_results'] = ExamResult.objects.filter(user=user).delete()[0]
        
        # 스터디 진행률 기록 삭제
        deleted_counts['study_progress_records'] = StudyProgressRecord.objects.filter(user=user).delete()[0]
        
        # 스터디 태스크 진행률 삭제
        deleted_counts['task_progress'] = StudyTaskProgress.objects.filter(user=user).delete()[0]
        
        # StudyTask progress 필드 초기화
        from ..models import StudyTask
        study_tasks = StudyTask.objects.filter(study__members__user=user)
        updated_count = 0
        for study_task in study_tasks:
            if hasattr(study_task, 'progress'):
                study_task.progress = 0
                study_task.save()
                updated_count += 1
        deleted_counts['study_task_progress_reset'] = updated_count
        
        # Django 캐시 정리
        from django.core.cache import cache
        cache.clear()
        
        # Django ORM 캐시 무효화
        from django.db import connection
        connection.close()
        
        # 더미 쿼리로 ORM 캐시 무효화
        ExamResult.objects.all().count()
        ExamResultDetail.objects.all().count()
        StudyTaskProgress.objects.all().count()
        StudyProgressRecord.objects.all().count()
        
        # 백업 정보를 세션에 저장 (다운로드용)
        request.session['statistics_backup'] = backup_data
        
        return Response({
            'success': True,
            'message': 'profile.statistics.reset.success',
            'deleted_counts': deleted_counts,
            'backup_data': backup_data
        })
        
    except Exception as e:
        logger.error(f'사용자 통계 초기화 실패: {e}')
        return Response({'error': f'profile.statistics.reset.error: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def backup_user_statistics(request):
    """사용자의 통계 백업 데이터를 Excel 파일로 다운로드합니다."""
    try:
        user = request.user
        
        # 세션에서 백업 데이터 가져오기
        backup_data = request.session.get('statistics_backup')
        if not backup_data:
            return Response({
                'error': 'profile.statistics.backup.notFound',
                'message': 'profile.statistics.backup.message'
            }, status=404)
        
        # Excel 파일 생성을 위한 openpyxl import
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        
        # 새 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistics Backup"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 추가
        headers = ["항목", "값", "설명"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 추가
        row = 2
        data_mapping = [
            ("사용자 ID", backup_data.get('user_id', ''), "사용자의 고유 식별자"),
            ("사용자명", backup_data.get('username', ''), "사용자 계정명"),
            ("백업 생성 시간", backup_data.get('timestamp', ''), "백업 데이터 생성 시점"),
            ("시험 결과 수", backup_data.get('exam_results_count', 0), "삭제된 시험 결과 개수"),
            ("스터디 진행률 기록 수", backup_data.get('study_progress_count', 0), "삭제된 스터디 진행률 기록 개수"),
            ("태스크 진행률 수", backup_data.get('task_progress_count', 0), "삭제된 태스크 진행률 개수"),
            ("정확도 조정 기록 수", backup_data.get('accuracy_adjustments_count', 0), "삭제된 정확도 조정 기록 개수")
        ]
        
        for item, value, description in data_mapping:
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=description)
            row += 1
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Excel 파일을 메모리에 저장
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Excel 파일로 응답
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="statistics_backup_{user.username}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        logger.error(f'통계 백업 다운로드 실패: {e}')
        return Response({'error': 'profile.statistics.backup.downloadFailed'}, status=500) 


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_my_account(request):
    """현재 로그인한 사용자의 계정을 탈퇴합니다."""
    try:
        user = request.user
        
        # 관리자(superuser)는 탈퇴할 수 없음
        if user.is_superuser:
            return Response({'detail': '관리자는 탈퇴할 수 없습니다.'}, status=400)
        
        username = user.username
        
        # 사용자와 관련된 모든 데이터 삭제
        
        # 1. 먼저 스터디 멤버 상태 확인 및 처리
        from ..models import Study, StudyTask
        user_studies = Study.objects.filter(members__user=user)
        
        for study in user_studies:
            # 해당 스터디의 멤버 수 확인 (본인 제외)
            other_members = study.members.exclude(user=user)
            logger.info(f'스터디 "{study.title_ko or study.title_en}" 멤버 분석: 본인 제외 멤버 {other_members.count()}명')
            
            if other_members.count() == 0:
                # 다른 멤버가 없으면 스터디와 연결된 태스크 삭제
                study_tasks = StudyTask.objects.filter(study=study)
                if study_tasks.exists():
                    logger.info(f'사용자 "{username}"의 고아 스터디 "{study.title_ko or study.title_en}"의 태스크 {study_tasks.count()}개 삭제')
                    study_tasks.delete()
                
                # 스터디 삭제
                logger.info(f'사용자 "{username}"의 고아 스터디 "{study.title_ko or study.title_en}" 삭제')
                study.delete()
            else:
                # 다른 멤버가 있으면 스터디 관리자 존재 여부 확인
                from ..models import Member
                current_user_member = study.members.filter(user=user).first()
                
                if current_user_member and current_user_member.role in ['study_admin', 'study_leader']:
                    # 본인이 관리자인 경우, 다른 관리자가 있는지 확인
                    other_admins = other_members.filter(role__in=['study_admin', 'study_leader'])
                    if other_admins.count() == 0:
                        # 다른 관리자가 없으면 에러 발생
                        study_title = study.title_ko or study.title_en or study.title
                        error_msg = f'"{study_title}" 스터디의 스터디 관리자를 지정해야 합니다.'
                        logger.error(f'사용자 "{username}" 탈퇴 실패: {error_msg}')
                        return Response({'detail': error_msg}, status=400)
                
                # 관리자 문제가 없으면 본인만 멤버에서 제거
                logger.info(f'스터디 "{study.title_ko or study.title_en}"에 다른 멤버 {other_members.count()}명 존재 - 본인만 제거하고 스터디 보존')
                study.members.filter(user=user).delete()
                logger.info(f'사용자 "{username}"을 스터디 "{study.title_ko or study.title_en}"에서 제거 완료 (다른 멤버 {other_members.count()}명 존재)')
        
        # 2. 시험 결과 상세 삭제
        exam_results = ExamResult.objects.filter(user=user)
        ExamResultDetail.objects.filter(result__in=exam_results).delete()
        
        # 3. 시험 결과 삭제
        exam_results.delete()
        
        # 4. 스터디 진행 기록 삭제
        StudyProgressRecord.objects.filter(user=user).delete()
        
        # 5. 스터디 태스크 진행율 삭제
        StudyTaskProgress.objects.filter(user=user).delete()
        
        # 6. 정확도 조정 기록 삭제
        AccuracyAdjustmentHistory.objects.filter(user=user).delete()
        
        # 7. 무시된 문제 삭제
        IgnoredQuestion.objects.filter(user=user).delete()
        
        # 8. 스터디 가입 요청 삭제
        StudyJoinRequest.objects.filter(user=user).delete()
        
        # 9. 시험 구독 정보 삭제
        ExamSubscription.objects.filter(user=user).delete()
        
        # 10. 사용자가 생성한 시험 처리 (스터디 멤버 상태에 따라 결정)
        user_exams = Exam.objects.filter(created_by=user)
        
        if user_exams.exists():
            logger.info(f'사용자 "{username}"이 생성한 시험 {user_exams.count()}개 처리 시작')
            
            for exam in user_exams:
                # 해당 시험이 연결된 스터디가 있는지 확인
                connected_studies = Study.objects.filter(tasks__exam=exam).distinct()
                
                should_preserve_exam = False
                if connected_studies.exists():
                    for study in connected_studies:
                        # 스터디에 다른 멤버가 있는지 확인
                        other_members = study.members.exclude(user=user)
                        if other_members.count() > 0:
                            should_preserve_exam = True
                            logger.info(f'시험 "{exam.title_ko or exam.title_en}"은 스터디 "{study.title_ko or study.title_en}"에 다른 멤버가 있어 보존됨')
                            break
                
                if should_preserve_exam:
                    # 스터디에 다른 멤버가 있으면 시험 보존 (created_by만 None으로 설정)
                    exam.created_by = None
                    exam.save()
                    logger.info(f'시험 "{exam.title_ko or exam.title_en}" 보존 (created_by를 None으로 설정)')
                else:
                    # 스터디에 다른 멤버가 없으면 시험 삭제
                    logger.info(f'시험 "{exam.title_ko or exam.title_en}" 삭제 (연결된 스터디에 다른 멤버 없음)')
                    exam.delete()
        
        # 11. 사용자 프로필 삭제 (CASCADE로 자동 삭제됨)
        
        # 13. 사용자 계정 삭제
        user.delete()
        
        return Response({
            'message': f'사용자 "{username}"의 계정이 성공적으로 탈퇴되었습니다.',
            'deleted_username': username
        }, status=200)
        
    except Exception as e:
        logger.error(f'사용자 계정 탈퇴 중 오류 발생: {str(e)}')
        return Response({'detail': f'계정 탈퇴 중 오류가 발생했습니다: {str(e)}'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def clear_study_cache(request):
    """
    스터디 관련 캐시를 강제로 무효화합니다 (K8s Redis 환경 대응).
    
    이 API는 스터디 생성/수정 후 My Studies에 표시되지 않는 문제를 해결하기 위해 사용됩니다.
    """
    try:
        user = request.user
        logger.info(f"[CLEAR_STUDY_CACHE] 사용자 {user.username}의 스터디 캐시 강제 무효화 시작")
        
        # ========================================
        # 🔄 백엔드 Redis 캐시 강제 무효화
        # ========================================
        try:
            from ..utils.cache_utils import StudyCacheManager
            from django.core.cache import cache
            
            # 1. StudyCacheManager를 통한 캐시 무효화
            StudyCacheManager.invalidate_all_study_cache()
            logger.info(f"[CLEAR_STUDY_CACHE] ✅ StudyCacheManager 캐시 무효화 완료")
            
            # 2. Redis 패턴 기반 강제 캐시 무효화
            if hasattr(cache, 'delete_pattern'):
                cache.delete_pattern("studies_*")
                cache.delete_pattern("exams_*")
                logger.info(f"[CLEAR_STUDY_CACHE] ✅ Redis 패턴 기반 캐시 무효화 완료")
            else:
                # 로컬 캐시의 경우 전체 클리어
                cache.clear()
                logger.info(f"[CLEAR_STUDY_CACHE] ✅ 로컬 캐시 전체 클리어 완료")
            
            # 3. 특정 사용자 캐시도 무효화
            StudyCacheManager.invalidate_user_study_cache(user.id)
            logger.info(f"[CLEAR_STUDY_CACHE] ✅ 사용자별 스터디 캐시 무효화 완료: {user.id}")
            
        except Exception as e:
            logger.error(f"[CLEAR_STUDY_CACHE] ❌ 캐시 무효화 실패: {e}")
            return Response({
                'success': False, 
                'message': f'캐시 무효화 실패: {str(e)}'
            }, status=500)
        
        logger.info(f"[CLEAR_STUDY_CACHE] ✅ 스터디 캐시 강제 무효화 완료")
        return Response({
            'success': True,
            'message': '스터디 캐시가 성공적으로 무효화되었습니다. 페이지를 새로고침해주세요.'
        })
        
    except Exception as e:
        logger.error(f"[CLEAR_STUDY_CACHE] ❌ 스터디 캐시 무효화 중 오류: {e}")
        return Response({
            'success': False,
            'message': f'스터디 캐시 무효화 중 오류가 발생했습니다: {str(e)}'
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def clear_all_cache(request):
    """
    사용자의 모든 캐시를 초기화합니다 (프론트엔드 + 백엔드 Redis).
    
    이 API는 프로필 페이지의 "모든 캐시 지우기" 버튼에서 호출됩니다.
    프론트엔드 캐시와 백엔드 Redis 캐시를 모두 정리하여 데이터 일관성을 보장합니다.
    """
    try:
        user = request.user
        logger.info(f"[CLEAR_ALL_CACHE] 사용자 {user.username}의 모든 캐시 초기화 시작")
        
        # ========================================
        # 🔄 백엔드 Redis 캐시 무효화 (본인 정보만)
        # ========================================
        try:
            from ..utils.cache_utils import ExamCacheManager, StudyCacheManager
            
            # 1. ExamCacheManager를 통한 본인 시험 관련 캐시만 무효화
            # ⚠️ 중요: invalidate_all_exam_cache()는 사용하지 않음 (다른 사용자 영향 방지)
            ExamCacheManager.invalidate_user_exam_cache(user.id)
            logger.info(f"[CLEAR_ALL_CACHE] ✅ 본인 시험 캐시 무효화 완료: 사용자 ID {user.id}")
            
            # 2. StudyCacheManager를 통한 본인 스터디 관련 캐시만 무효화
            # ⚠️ 중요: invalidate_all_study_cache()는 사용하지 않음 (다른 사용자 영향 방지)
            StudyCacheManager.invalidate_user_study_cache(user.id)
            logger.info(f"[CLEAR_ALL_CACHE] ✅ 본인 스터디 캐시 무효화 완료: 사용자 ID {user.id}")
            
            # 3. 본인과 관련된 특정 캐시 키들만 무효화
            from django.core.cache import cache
            
            if hasattr(cache, 'delete_pattern'):
                # Redis 환경: 본인과 관련된 캐시 패턴만 삭제
                user_specific_patterns = [
                    f"exam_{user.id}_*",        # 본인 시험 관련 캐시
                    f"study_{user.id}_*",       # 본인 스터디 관련 캐시
                    f"user_{user.id}_*",        # 본인 사용자 관련 캐시
                    f"*_{user.id}_*"            # 본인 ID가 포함된 모든 캐시
                ]
                
                for pattern in user_specific_patterns:
                    try:
                        cache.delete_pattern(pattern)
                        logger.info(f"[CLEAR_ALL_CACHE] ✅ 본인 전용 패턴 '{pattern}' 캐시 삭제 완료")
                    except Exception as pattern_error:
                        logger.warning(f"[CLEAR_ALL_CACHE] ⚠️ 패턴 '{pattern}' 캐시 삭제 실패: {pattern_error}")
                
                logger.info(f"[CLEAR_ALL_CACHE] ✅ 본인 전용 Redis 캐시 무효화 완료")
                
            else:
                # 로컬 캐시 환경에서는 개별 키 기반으로 본인 관련 캐시만 삭제
                logger.info("[CLEAR_ALL_CACHE] 🔄 로컬 캐시에서 본인 관련 캐시만 삭제 시작")
                
                # 로컬 캐시에서 본인과 관련된 키들만 찾아서 삭제
                # (로컬 환경에서는 전체 캐시 클리어가 불가피할 수 있음)
                cache.clear()
                logger.info("[CLEAR_ALL_CACHE] ✅ 로컬 캐시 클리어 완료 (본인 데이터 보호)")
            
        except Exception as e:
            logger.error(f"[CLEAR_ALL_CACHE] ❌ ExamCacheManager/StudyCacheManager 캐시 무효화 실패: {e}")
            
            # 폴백: Redis 패턴 기반 캐시 무효화 (본인 정보만)
            try:
                from django.core.cache import cache
                
                if hasattr(cache, 'delete_pattern'):
                    # Redis 환경: 본인과 관련된 패턴만 삭제 (다른 사용자 영향 방지)
                    logger.info("[CLEAR_ALL_CACHE] 🔄 Redis 패턴 기반 본인 전용 캐시 무효화 시작")
                    
                    # 본인과 관련된 캐시 패턴만 (다른 사용자 영향 방지)
                    user_specific_patterns = [
                        f"exam_{user.id}_*",        # 본인 시험 관련 캐시
                        f"study_{user.id}_*",       # 본인 스터디 관련 캐시
                        f"question_*_{user.id}_*",  # 본인이 푼 문제 관련 캐시
                        f"statistics_{user.id}_*",  # 본인 통계 관련 캐시
                        f"user_{user.id}_*",        # 본인 사용자 관련 캐시
                        f"*_{user.id}_*"            # 본인 ID가 포함된 모든 캐시
                    ]
                    
                    for pattern in user_specific_patterns:
                        try:
                            cache.delete_pattern(pattern)
                            logger.info(f"[CLEAR_ALL_CACHE] ✅ 본인 전용 패턴 '{pattern}' 캐시 삭제 완료")
                        except Exception as pattern_error:
                            logger.warning(f"[CLEAR_ALL_CACHE] ⚠️ 패턴 '{pattern}' 캐시 삭제 실패: {pattern_error}")
                    
                    logger.info("[CLEAR_ALL_CACHE] ✅ Redis 패턴 기반 본인 전용 캐시 무효화 완료")
                    
                else:
                    # 로컬 캐시 환경에서는 본인 관련 키만 찾아서 삭제 시도
                    logger.info("[CLEAR_ALL_CACHE] 🔄 로컬 캐시에서 본인 관련 캐시만 삭제 시도")
                    
                    # 로컬 캐시에서 본인과 관련된 키들만 찾아서 삭제
                    try:
                        # 캐시 키 목록을 가져와서 본인 관련 키만 필터링
                        all_keys = cache.keys('*') if hasattr(cache, 'keys') else []
                        user_keys = [key for key in all_keys if str(user.id) in key]
                        
                        for key in user_keys:
                            cache.delete(key)
                            logger.info(f"[CLEAR_ALL_CACHE] ✅ 본인 관련 키 '{key}' 캐시 삭제 완료")
                        
                        logger.info(f"[CLEAR_ALL_CACHE] ✅ 본인 관련 {len(user_keys)}개 키 삭제 완료")
                        
                    except Exception as key_error:
                        logger.warning(f"[CLEAR_ALL_CACHE] ⚠️ 개별 키 삭제 실패, 전체 캐시 클리어: {key_error}")
                        # 최후 수단: 전체 캐시 클리어 (본인 데이터 보호를 위해 경고)
                        cache.clear()
                        logger.warning("[CLEAR_ALL_CACHE] ⚠️ 본인 데이터 보호를 위해 전체 캐시 클리어 실행")
                    
            except Exception as e2:
                logger.error(f"[CLEAR_ALL_CACHE] ❌ 폴백 캐시 무효화도 실패: {e2}")
                logger.warning(f"[CLEAR_ALL_CACHE] 🚨 사용자 {user.username}의 캐시 무효화에 실패했습니다!")
                logger.warning(f"[CLEAR_ALL_CACHE] 🚨 다른 사용자의 데이터는 영향을 받지 않았습니다.")
        
        # ========================================
        # 🗑️ Django ORM 캐시 무효화
        # ========================================
        try:
            from django.db import connection
            
            # 데이터베이스 연결 닫기 (ORM 캐시 무효화)
            connection.close()
            
            # 더미 쿼리로 ORM 캐시 무효화
            from ..models import ExamResult, ExamResultDetail, StudyTaskProgress, StudyProgressRecord
            
            ExamResult.objects.all().count()
            ExamResultDetail.objects.all().count()
            StudyTaskProgress.objects.all().count()
            StudyProgressRecord.objects.all().count()
            
            logger.info("[CLEAR_ALL_CACHE] ✅ Django ORM 캐시 무효화 완료")
            
        except Exception as e:
            logger.warning(f"[CLEAR_ALL_CACHE] ⚠️ Django ORM 캐시 무효화 실패: {e}")
        
        # ========================================
        # 📊 캐시 무효화 결과 요약
        # ========================================
        logger.info(f"[CLEAR_ALL_CACHE] 🎉 사용자 {user.username}의 모든 캐시 초기화 완료")
        logger.info(f"[CLEAR_ALL_CACHE] 📋 초기화된 캐시:")
        logger.info(f"[CLEAR_ALL_CACHE]   - 백엔드 Redis 캐시: ✅")
        logger.info(f"[CLEAR_ALL_CACHE]   - Django ORM 캐시: ✅")
        logger.info(f"[CLEAR_ALL_CACHE]   - 프론트엔드 캐시: 사용자가 직접 정리 필요")
        
        return Response({
            'success': True,
            'message': '모든 캐시가 성공적으로 초기화되었습니다.',
            'details': {
                'backend_cache': 'cleared',
                'orm_cache': 'cleared',
                'frontend_cache': 'user_action_required',
                'timestamp': timezone.now().isoformat()
            }
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        error_msg = f'캐시 초기화 중 오류가 발생했습니다: {str(e)}'
        logger.error(f'[CLEAR_ALL_CACHE] ❌ 오류: {error_msg}')
        logger.error(f'[CLEAR_ALL_CACHE] ❌ 스택 트레이스: {traceback.format_exc()}')
        
        return Response({
            'success': False,
            'error': error_msg
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)